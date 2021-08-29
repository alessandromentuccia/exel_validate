import argparse
import itertools
import json
import logging
import random
import re
import time
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List

import openpyxl 
import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import xlsxwriter
#from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)


class Validator_v():

    file_name = ""
    file_data = {}
    catalogo = OrderedDict()


    def ck_QD_description(self, df_mapping, sheet_QD):
        print("start checking if foreach agenda there are the same QD")
        
        error_dict = {
            'vuoti_description_QD': [],
            'disciplina_non_corretta': []
            }
        
        xfile = openpyxl.load_workbook(self.file_name)
        sheet = xfile.get_sheet_by_name('Tracciato Mapping') #modificare sheet name

        for index, row in df_mapping.iterrows(): 
            QD_list = row["Codice Quesito Diagnostico"].split(",") #lista dei codici QD
            disciplina = str(row["Codici disciplina da catalogo"]).strip() #codice disciplina

            if QD_list != "":
                string_record = []
                disciplina_record = [] #QD che non hanno la disciplina corretta
                QD_record = [] #QD non trovati e da riportare in error
                flag_error_disciplina = False
                flag_error_QD = False
                cell_QD_value = Validator.findCell_col_all(self, sheet_QD, disciplina, 0, 1) #[] 

                for QD in QD_list: #ciclo per ogni QD
                    if QD != "":
                        QD = QD.strip() #elimino eventuali spazi ai bordi
                        rowQD, colQD = Validator.findCell_col(self, sheet_QD, QD, 5, 6)

                        if cell_QD_value != []:
                            if QD not in cell_QD_value:
                                print("QD: " + QD + ", cell_QD_value: %s", cell_QD_value)
                                disciplina_record.append(QD)   
                                flag_error_disciplina = True
                        
                                
                        if rowQD != -1:
                            #resultQDdisciplina = sheet_QD.cell_value(rowQD, 0) #valore disciplina QD da catalogo
                            resultQD_description = sheet_QD.cell_value(rowQD, 6) #valore descrizione QD da catalogo
                            string_record.append(resultQD_description)
                        if rowQD == -1 or Validator.findCell(self, sheet_QD, QD, 5, 6) == -1:
                            print("vuoti_description_QD: " + QD)
                            QD_record.append(QD)
                            flag_error_QD = True


                sheet["AH"+str(int(index)+2)] = ",".join(string_record) #modificare colonna QD descrizione
                print("descrizione aggiunta")

                st_error = ""
                if flag_error_disciplina == True: #controllo se c'è stato errore di desciplina nella riga
                    st_error = "alert: disciplina non corrispondente per QD: " + ",".join(disciplina_record) + "\n"
                    print("disciplina errore aggiunta")
                    error_dict["disciplina_non_corretta"].append(str(int(index)+2))

                #disciplina_res = self.check_disciplina_per_record(QD_list, sheet_QD)
                #if disciplina_res != -1:
                #    sheet["A"+str(int(index)+2)] = ",".join(disciplina_res)
                #elif disciplina_res == -1:
                #    st_error = "alert: Non esiste una disciplina in comune tra i QD all'indice: " + str(int(index)+2) + "\n"

                if flag_error_QD == True:
                    st_error =  st_error + "alert: QD non presente nel catalogo: " + ",".join(QD_record)
                    error_dict["vuoti_description_QD"].append(str(int(index)+2))
                
                if st_error != "":
                    sheet["BP"+str(int(index)+2)] = st_error #modificare colonna alert

        xfile.save(self.file_name)
        return error_dict


    def ck_metodiche_description(self, df_mapping, sheet_Metodiche):
        print("start checking metodiche")
        
        error_dict = {
            'metodiche_non_in_prestazione': [],
            'metodica non_esiste': []
            }
        
        xfile = openpyxl.load_workbook(self.file_name)
        sheet = xfile.get_sheet_by_name('Tracciato Mapping') #modificare sheet name

        for index, row in df_mapping.iterrows(): 
            metodica_list = row["Codice Metodica"].split(",") #lista dei codici QD
            cod_siss =  row["Codice Prestazione SISS"]

            if metodica_list != "":
                string_record = []
                prestazione_record = [] #QD che non hanno la disciplina corretta
                metodica_record = [] #QD non trovati e da riportare in error
                flag_error_prestazione = False
                flag_error_metodica = False
                cell_metodica_value = Validator.findCell_all_SISS(self, sheet_Metodiche, cod_siss, 2, 3) #[] 

                for metodica in metodica_list: #ciclo per ogni QD
                    if metodica != "":
                        metodica = metodica.strip() #elimino eventuali spazi ai bordi
                        rowmetodica, colmetodica = Validator.findCell_col(self, sheet_Metodiche, metodica, 4, 5)

                        if cell_metodica_value != []:
                            if metodica not in cell_metodica_value:
                                print("metodica: " + metodica + ", cell_metodica_value: %s", cell_metodica_value)
                                prestazione_record.append(metodica)   
                                flag_error_prestazione = True
                        
                                
                        if rowmetodica != -1:
                            #resultQDdisciplina = sheet_QD.cell_value(rowQD, 0) #valore disciplina QD da catalogo
                            resultmetodica_description = sheet_Metodiche.cell_value(rowmetodica, 5) #valore descrizione QD da catalogo
                            string_record.append(resultmetodica_description)
                        if rowmetodica == -1 or Validator.findCell(self, sheet_Metodiche, metodica, 4, 5) == -1:
                            print("metodica non_esiste: " + metodica)
                            metodica_record.append(metodica)
                            flag_error_metodica = True

                        
                sheet["AN"+str(int(index)+2)] = ",".join(string_record) #modificare colonna QD descrizione
                print("descrizione aggiunta")

                st_error = ""
                if flag_error_prestazione == True: #controllo se c'è stato errore di desciplina nella riga
                    st_error = "alert: codice prestazione SISS non corrispondente per metodica: " + ",".join(prestazione_record) + "\n"
                    print("errore relazione metodica codice SISS")
                    error_dict["metodiche_non_in_prestazione"].append(str(int(index)+2))

                #disciplina_res = self.check_disciplina_per_record(QD_list, sheet_QD)
                #if disciplina_res != -1:
                #    sheet["A"+str(int(index)+2)] = ",".join(disciplina_res)
                #elif disciplina_res == -1:
                #    st_error = "alert: Non esiste una disciplina in comune tra i QD all'indice: " + str(int(index)+2) + "\n"

                if flag_error_metodica == True:
                    st_error =  st_error + "alert: metodica non presente nel catalogo: " + ",".join(metodica_record)
                    error_dict["metodica non_esiste"].append(str(int(index)+2))
                
                if st_error != "":
                    sheet["BQ"+str(int(index)+2)] = st_error #modificare colonna alert

        xfile.save(self.file_name)
        return error_dict

    
    def ck_distretti_description(self, df_mapping, sheet_Distretti):
        print("start checking distretti")
        
        error_dict = {
            'distretti_non_in_prestazione': [],
            'distretti non_esiste': []
            }
        
        xfile = openpyxl.load_workbook(self.file_name)
        sheet = xfile.get_sheet_by_name('Tracciato Mapping') #modificare sheet name

        for index, row in df_mapping.iterrows(): 
            metodica_list = row["Codice Metodica"].split(",") #lista dei codici QD
            cod_siss =  row["Codice Prestazione SISS"]

            if metodica_list != "":
                string_record = []
                prestazione_record = [] #QD che non hanno la disciplina corretta
                metodica_record = [] #QD non trovati e da riportare in error
                flag_error_prestazione = False
                flag_error_metodica = False
                cell_metodica_value = Validator.findCell_all_SISS(self, sheet_Distretti, cod_siss, 2, 3) #[] 

                for metodica in metodica_list: #ciclo per ogni QD
                    if metodica != "":
                        metodica = metodica.strip() #elimino eventuali spazi ai bordi
                        rowmetodica, colmetodica = Validator.findCell_col(self, sheet_Distretti, metodica, 4, 5)

                        if cell_metodica_value != []:
                            if metodica not in cell_metodica_value:
                                print("metodica: " + metodica + ", cell_metodica_value: %s", cell_metodica_value)
                                prestazione_record.append(metodica)   
                                flag_error_prestazione = True
                        
                                
                        if rowmetodica != -1:
                            #resultQDdisciplina = sheet_QD.cell_value(rowQD, 0) #valore disciplina QD da catalogo
                            resultmetodica_description = sheet_Distretti.cell_value(rowmetodica, 5) #valore descrizione QD da catalogo
                            string_record.append(resultmetodica_description)
                        if rowmetodica == -1 or Validator.findCell(self, sheet_Distretti, metodica, 4, 5) == -1:
                            print("metodica non_esiste: " + metodica)
                            metodica_record.append(metodica)
                            flag_error_metodica = True

                        
                sheet["AN"+str(int(index)+2)] = ",".join(string_record) #modificare colonna QD descrizione
                print("descrizione aggiunta")

                st_error = ""
                if flag_error_prestazione == True: #controllo se c'è stato errore di desciplina nella riga
                    st_error = "alert: codice prestazione SISS non corrispondente per metodica: " + ",".join(prestazione_record) + "\n"
                    print("errore relazione metodica codice SISS")
                    error_dict["metodiche_non_in_prestazione"].append(str(int(index)+2))

                #disciplina_res = self.check_disciplina_per_record(QD_list, sheet_QD)
                #if disciplina_res != -1:
                #    sheet["A"+str(int(index)+2)] = ",".join(disciplina_res)
                #elif disciplina_res == -1:
                #    st_error = "alert: Non esiste una disciplina in comune tra i QD all'indice: " + str(int(index)+2) + "\n"

                if flag_error_metodica == True:
                    st_error =  st_error + "alert: metodica non presente nel catalogo: " + ",".join(metodica_record)
                    error_dict["metodica non_esiste"].append(str(int(index)+2))
                
                if st_error != "":
                    sheet["BQ"+str(int(index)+2)] = st_error #modificare colonna alert

        xfile.save(self.file_name)
        return error_dict




    def findCell(self, sh, searchedValue, start_col, end_col):
        for row in range(sh.nrows):
            for col in range(start_col, end_col):
                myCell = sh.cell(row, col)
                if myCell.value == searchedValue:
                    return row, col#xl_rowcol_to_cell(row, col)
        return -1, -1

    def findCell_col(self, sh, searchedValue, start_col, end_col):
        for row in range(sh.nrows):
            for col in range(start_col, end_col):
                myCell = sh.cell(row, col)
                if myCell.value == searchedValue:
                    return row, col#xl_rowcol_to_cell(row, col)
        return -1, -1

    def findCell_col_all(self, sh, searchedValue, start_col, end_col):
        result = []
        cell_value = []
        QD_value = []
        for row in range(sh.nrows):
            for col in range(start_col, end_col):
                myCell = sh.cell(row, col).value # celle trovate
                QD_cell = sh.cell(row, col+5).value #
                #print("tipo disciplina catalogo: " + str(type(QD_cell)))
                if str(searchedValue) == str(myCell) and QD_cell not in QD_value:
                    #    print("searchedValue: " + searchedValue + ", myCell.value: " + QD_cell.value)
                    result.append(str(row) + "#" + str(col)) #return row, col#xl_rowcol_to_cell(row, col)
                    cell_value.append(myCell) 
                    QD_value.append(QD_cell)    
        return QD_value

    def findCell_all_SISS(self, sh, searchedValue, start_col, end_col):
        result = []
        cell_value = []
        metodica_value = []
        for row in range(sh.nrows):
            for col in range(start_col, end_col):
                myCell = sh.cell(row, col).value # SISS
                metodica_cell = sh.cell(row, col+2).value # cod metodica
                #print("tipo disciplina catalogo: " + str(type(QD_cell)))
                if str(searchedValue) == str(myCell) and metodica_cell not in metodica_value:
                    #    print("searchedValue: " + searchedValue + ", myCell.value: " + QD_cell.value)
                    result.append(str(row) + "#" + str(col)) #return row, col#xl_rowcol_to_cell(row, col)
                    cell_value.append(myCell) 
                    metodica_value.append(metodica_cell)    
        return metodica_value

    def findCell_discipline_per_QD(self, sh, searchedValue, start_col, end_col):
        result = []
        cell_value = []
        QD_value = []
        for row in range(sh.nrows):
            for col in range(start_col, end_col):
                myCell = sh.cell(row, col).value # celle trovate
                Dish_cell = sh.cell(row, 0).value #
                if str(searchedValue) == str(myCell) and Dish_cell not in QD_value:
                    print("QD: " + myCell + " disciplina: " + str(Dish_cell))
                    #if QD_cell == "QD00146" or QD_cell == "QD00948":
                    #    print("dona errore")
                    #    print("searchedValue: " + searchedValue + ", myCell.value: " + QD_cell.value)
                    #result.append(str(row) + "#" + str(col)) #return row, col#xl_rowcol_to_cell(row, col)
                    cell_value.append(myCell) 
                    QD_value.append(Dish_cell)    
        return QD_value

    def findCell_lasco(self, sh, searchedValue, start_col):
        result = []
        cell_value = []
        for row in range(sh.nrows):
            for col in range(start_col, sh.ncols):
                myCell = sh.cell(row, col)
                if searchedValue in myCell.value and myCell.value not in cell_value:
                    result.append(str(row) + "#" + str(col)) #return row, col#xl_rowcol_to_cell(row, col)
                    cell_value.append(myCell.value) 
                    #print("searchedValue: " + searchedValue + ", myCell.value: " + myCell.value)
                    #print("row: " + str(row) + "col: " + str(col))
        if result == []:
            return -1
        return result

    


