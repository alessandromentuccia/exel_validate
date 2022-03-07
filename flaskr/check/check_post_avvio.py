import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import yaml
import logging
import re
import openpyxl 

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

class Check_post_avvio():
    
    def ck_post_avvio(self, df_mapping, df_rivisto, error_dict):
        print("start checking post avvio")
        error_dict.update({'error_post_avvio': []})

        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel
 
        if self.configurazione_rivisto["Quesiti"] != "": #check list
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Quesiti", self.work_codice_QD)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Quesiti", self.work_codice_QD)
            #error_dict = Check_post_avvio.ck_QD(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["OperatoreQD"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "OperatoreQD", self.work_operatore_logico_QD)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "OperatoreQD", self.work_operatore_logico_QD)
            #error_dict = Check_post_avvio.ck_Operatore_QD(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Distretti"] != "": #check list
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Distretti", self.work_codice_distretto)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Distretti", self.work_codice_distretto)
            #error_dict = Check_post_avvio.ck_Distretti(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["OperatoreDistretto"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "OperatoreDistretto", self.work_operatore_logico_distretto)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "OperatoreDistretto", self.work_operatore_logico_distretto)
            #error_dict = Check_post_avvio.ck_Operatore_Distretti(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Metodiche"] != "": #check list
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Metodiche", self.work_codice_metodica)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Metodiche", self.work_codice_metodica)
            #error_dict = Check_post_avvio.ck_Metodiche(self, df_mapping, df_rivisto, error_dict)    
        if self.configurazione_rivisto["Inviante"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Inviante", self.work_inviante)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Inviante", self.work_inviante)
            #error_dict = Check_post_avvio.ck_Invianti(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Risorsa"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Risorsa", self.work_risorsa)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Risorsa", self.work_risorsa)
            #error_dict = Check_post_avvio.ck_Risorsa(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Farmacia"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Farmacia", self.work_accesso_farmacia)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Farmacia", self.work_accesso_farmacia)
            #error_dict = Check_post_avvio.ck_Farmacia(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["CCR"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "CCR", self.work_accesso_CCR)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "CCR", self.work_accesso_CCR)
            #error_dict = Check_post_avvio.ck_CCR(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Cittadino"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Cittadino", self.work_accesso_cittadino)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Cittadino", self.work_accesso_cittadino)
            #error_dict = Check_post_avvio.ck_Cittadino(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["MMG"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "MMG", self.work_accesso_MMG)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "MMG", self.work_accesso_MMG)
            #error_dict = Check_post_avvio.ck_MMG(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Amministrativo"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Amministrativo", self.work_accesso_amministrativo)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Amministrativo", self.work_accesso_amministrativo)
            #error_dict = Check_post_avvio.ck_Amministrativo(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["PAI"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "PAI", self.work_accesso_PAI)
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "PAI", self.work_accesso_PAI)
            #error_dict = Check_post_avvio.ck_PAI(self, df_mapping, df_rivisto, error_dict)

        #creazione del messaggio di alert riportato nel file excel
        print("start definizione output controlli post avvio")
        return error_dict

    '''metodo che controlla se gli attributi del rivisto sono configurati allo stesso modo anche nel mapping'''
    def ck_MAP(self, df_mapping, df_rivisto, error_dict, element, work_codice):
        print("start checking MAPPING")
        error_dict.update({
            'error_ck_'+element: [] })

        alert_column = "DP"

        xfile = openpyxl.load_workbook(self.file_rivisto) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.configurazione_rivisto["Sheet"])

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        agenda = configurazione_rivisto["Agenda"] #intesta
        prestazioneSISS = configurazione_rivisto["PrestazioneSISS"]
        prestazioneInterna = configurazione_rivisto["PrestazioneInterna"]

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[agenda]).strip()+"|"+str(row[prestazioneSISS]).strip()+"|"+str(row[prestazioneInterna]).strip()
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[configurazione_rivisto[element]] #modificare
            if element == "Inviante":
                if searchedValue == "":
                    searchedValue = "0"
            result_ = self.findCell_dataframe_MAP(df_mapping, searchedValue, rivisto_key, work_codice) #modificare
            if result_ == -1:
                error_dict["error_ck_"+element].append(str(int(index)+2))
                print("trovato errore su "+element)
                out_message = "__> {}".format("Corrispondenza "+element+" non trovata in mapping")
                if sheet[alert_column+str(int(index)+2)].value is not None:
                    sheet[alert_column+str(int(index)+2)] = str(sheet[alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[alert_column+str(int(index)+2)] = out_message
                
        xfile.save(self.file_rivisto)     
        return error_dict
    
    '''Metodo che controlla se gli attributi del mapping sono anche sul rivisto'''
    def ck_RIV(self, df_mapping, df_rivisto, error_dict, element, work_codice):
        print("start checking RIVISTO")
        error_dict.update({
            'error_ck_reverse_'+element: [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)
        
        file_rivisto = self.file_rivisto  
        configurazione_rivisto = self.configurazione_rivisto

        agenda = self.work_codice_agenda_siss #intesta
        prestazioneSISS = self.work_codice_prestazione_siss
        prestazioneInterna = self.work_codice_prestazione_interno

        for index, row in df_mapping.iterrows():
            if row[self.work_abilitazione_esposizione_siss] == "S":
                mapping_key = str(row[agenda]).strip()+"|"+str(row[prestazioneSISS]).strip()+"|"+str(row[prestazioneInterna]).strip()
                #print("iterate row: " + rivisto_key)
                #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
                searchedValue = row[work_codice] #modificare
                result_ = self.findCell_dataframe_RIV(df_rivisto, searchedValue, mapping_key, configurazione_rivisto[element]) #modificare
                if result_ == -1:
                    error_dict["error_ck_reverse_"+element].append(str(int(index)+2))
                    print("trovato errore su "+element)
                    out_message = "__> {}".format("Corrispondenza "+element+" non trovata in rivisto")
                    if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                        sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                    else:
                        sheet[self.work_alert_column+str(int(index)+2)] = out_message
                
        xfile.save(self.file_data)     
        return error_dict

    '''def ck_QD(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_QD': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Quesiti"]] #modificare
            result_QD = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, self.work_codice_QD) #modificare
            if result_QD == -1:
                error_dict["error_ck_QD"].append(str(int(index)+2))
                print("trovato errore su Quesiti")
                out_message = "__> {}".format("Corrispondenza di Quesiti non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message
                
        xfile.save(self.file_data)     
        return error_dict

    def ck_Operatore_QD(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Operatore_QD': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["OperatoreQD"]] #modificare
            result_operatore_QD = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, self.work_operatore_logico_QD) #modificare
            if result_operatore_QD == -1:
                error_dict["error_ck_Operatore_QD"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza dell'operatore logico QD non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_Distretti(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Distretti': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Distretti"]] #modificare
            result_Distretti = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, self.work_codice_distretto) #modificare
            if result_Distretti == -1:
                error_dict["error_ck_Distretti"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza dei distretti non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_Operatore_Distretti(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Operatore_Distretti': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["OperatoreDistretto"]] #modificare
            result_Operatore_Distretti = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, self.work_operatore_logico_distretto) #modificare
            if result_Operatore_Distretti == -1:
                error_dict["error_ck_Operatore_Distretti"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza dell'operatore logico distretti non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_Metodiche(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Metodiche': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Metodiche"]] #modificare
            result_Metodiche = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, self.work_codice_metodica) #modificare
            if result_Metodiche == -1:
                error_dict["error_ck_Metodiche"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza delle metodiche non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict
    
    def ck_Invianti(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Invianti': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Inviante"]] #modificare
            result_Invianti = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, self.work_inviante) #modificare
            if result_Invianti == -1:
                error_dict["error_ck_Invianti"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza dell'inviante non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_Risorsa(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Risorsa': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Risorsa"]] #modificare
            result_Risorsa = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, self.work_risorsa) #modificare
            if result_Risorsa == -1:
                error_dict["error_ck_Risorsa"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza della risorsa non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_Farmacia(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Farmacia': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        column_farmacia = self.work_accesso_farmacia

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Farmacia"]] #modificare
            
            result_farmacia = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, column_farmacia) 
            if result_farmacia == -1:
                error_dict["error_ck_Farmacia"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza canale farmacia non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict
    
    def ck_CCR(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_CCR': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        column_CCR = self.work_accesso_CCR

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["CCR"]] #modificare
            
            result_CCR = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, column_CCR) 
            if result_CCR == -1:
                error_dict["error_ck_CCR"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza canale CCR non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_Cittadino(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Cittadino': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        column_cittadino = self.work_accesso_cittadino

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Cittadino"]] #modificare
            
            result_cittadino = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, column_cittadino) 
            if result_cittadino == -1:
                error_dict["error_ck_Cittadino"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza canale cittadino non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_MMG(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Farmacia': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        column_MMG = self.work_accesso_MMG

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["MMG"]] #modificare
            
            result_MMG = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, column_MMG) 
            if result_MMG == -1:
                error_dict["error_ck_MMG"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza canale MMG non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_Amministrativo(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_Amministrativo': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        column_amministrativo = self.work_accesso_amministrativo

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["Amministrativo"]] #modificare
            
            result_amministrativo = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, column_amministrativo) 
            if result_amministrativo == -1:
                error_dict["error_ck_Amministrativo"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza canale amministrativo non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict

    def ck_PAI(self, df_mapping, df_rivisto, error_dict):
        error_dict.update({
            'error_ck_PAI': [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        column_PAI = self.work_accesso_PAI

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[self.configurazione_rivisto["Agenda"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneSISS"]].strip())+"|"+str(row[self.configurazione_rivisto["PrestazioneInterna"]].strip())
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[self.configurazione_rivisto["PAI"]] 
            
            result_PAI = self.findCell_dataframe(df_mapping, searchedValue, rivisto_key, column_PAI) 
            if result_PAI == -1:
                error_dict["error_ck_PAI"].append(str(int(index)+2))
                out_message = "__> {}".format("Corrispondenza canale PAI non trovata in mapping")
                if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                    sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[self.work_alert_column+str(int(index)+2)] = out_message

        xfile.save(self.file_data)     
        return error_dict'''
    