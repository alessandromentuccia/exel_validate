import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import yaml
import logging
import re
import openpyxl 
import sys
from pathlib import Path

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

class Check_QD():

    file_name = ""
    output_message = ""
    error_list = {}
    file_data = {}

    work_sheet = "" #sheet di lavoro di df_mapping
    work_codice_prestazione_siss = ""
    work_descrizione_prestazione_siss = ""
    work_codice_agenda_siss = ""
    work_casi_1_n = ""
    work_abilitazione_esposizione_siss = ""
    work_codici_disciplina_catalogo = ""
    work_descrizione_disciplina_catalogo = ""
    work_codice_QD = ""
    work_descrizione_QD = ""
    work_operatore_logico_QD = ""
    work_codice_metodica = ""
    work_descrizione_metodica = ""
    work_codice_distretto = ""
    work_descrizione_distretto = ""
    work_operatore_logico_distretto = ""
    work_priorita_U = ""
    work_priorita_primo_accesso_D = ""
    work_priorita_primo_accesso_P = ""
    work_priorita_primo_accesso_B = ""
    work_accesso_programmabile_ZP = ""

    work_index_codice_QD = 0
    work_index_codice_SISS_agenda = 0
    work_index_abilitazione_esposizione_SISS = 0
    work_index_codice_prestazione_SISS = 0
    work_index_operatore_logico_distretto = 0
    work_index_codici_disciplina_catalogo = 0
    work_index_codici_descri_disciplina_catalogo = 0


    def ck_QD_agenda(self, df_mapping, error_dict):
        print("start checking if foreach agenda there are the same QD")
        
        error_dict.update({
            'error_QD_agenda': [],
            'error_QD_vuoto' : []})
            
        QD_dict_error = {}
        index_agenda_dict = {}
        QD_agenda_dict = {}

        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel
        
        #agenda = df_mapping[self.work_codice_agenda_siss].iloc[2]
        #last_QD = df_mapping[self.work_codice_QD].iloc[2]
        for index, row in df_mapping.iterrows():
            if row[self.work_abilitazione_esposizione_siss] == "S":
                index_agenda_dict = self.update_list_in_dict(index_agenda_dict, row[self.work_codice_agenda_siss], str(int(index)+2))
                QD_agenda_dict[str(int(index)+2)] = str(row[self.work_codice_QD])
            
        for key, indice in index_agenda_dict.items(): #key: AGENDA, value: INDICE
            flag_QD_non_vuoto = False
            flag_QD_diverso = False

            QD_list_vuoti = []
            agenda = key
            cont = 0
            QD_list_last = []
            QD_list_last_last = []
            if len(indice) > 1:
                for ind in indice: 
                    QD_string = QD_agenda_dict[ind] #stringa dei QD
                    print("indice: " + ind + "QD_string: " + QD_string)
                    QD_prestazione = QD_string.split(self.work_delimiter) #lista dei QD
                    QD_lista_prestazioni = []
                    for QD in QD_prestazione:
                        QD_lista_prestazioni.append(QD.strip()) #lista dei QD ripulita dagli spazi
                    QD_lista_prestazioni.sort() #lista dei QD ordinata
                    if cont == 0:
                        cont = cont + 1                    
                        QD_list_last = QD_lista_prestazioni 
                    print("QD_lista_prestazioni", QD_lista_prestazioni)
                    if QD_lista_prestazioni == ['']:
                        QD_list_vuoti.append(ind)
                        #QD_list_last = QD_list_last_last
                    else:
                        flag_QD_non_vuoto = True
                        #D_list_last.sort()
                        if QD_lista_prestazioni != QD_list_last:
                            error_dict['error_QD_agenda'].append(ind)
                            print("trovato QD non coerente in agenda")
                            QD_dict_error = self.update_list_in_dict(QD_dict_error, ind, QD_string)
                        QD_list_last = QD_lista_prestazioni
                        #QD_list_last_last = QD_list_last
            
            if flag_QD_non_vuoto == True and QD_list_vuoti != ['']:
                for ind in QD_list_vuoti:
                    error_dict['error_QD_vuoto'].append(ind)
            

            '''if row[self.work_abilitazione_esposizione_siss] == "S":
                if row[self.work_codice_agenda_siss] == agenda:
                    if row[self.work_codice_QD] != last_QD:
                        print("error QD at index:" +  str(int(index)+2))
                        #error_list.append(str(int(index)+2))
                        error_dict['error_QD_agenda'].append(str(int(index)+2))
                        QD_dict_error = self.update_list_in_dict(QD_dict_error, str(int(index)+2), row[self.work_codice_QD])
                else: 
                    agenda = row[self.work_codice_agenda_siss]
                    last_QD = row[self.work_codice_QD]''' 
        print("start definizione output agenda QD")
        out1 = ""
        out_message = ""
        for ind in error_dict['error_QD_agenda']:
            out1 = out1 + "at index: " + ind + ", ".join(QD_dict_error[ind]) + ", \n"
            out_message = "__> QD: '{}' diversi per le prestazioni della stessa agenda".format(", ".join(QD_dict_error[ind]))
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        print("print output qd vuoto")        
        for ind in error_dict['error_QD_vuoto']:
            #out1 = out1 + "at index: " + ind + ", ".join(QD_dict_error[ind]) + ", \n"
            out_message = "__> QD: prestazione con QD vuoti nell'agenda"
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        self.output_message = self.output_message + "\nerror_QD_agenda: \n" + out1
        print("finish print output qd vuoto")  
        xfile.save(self.file_data) 
        return error_dict

    def ck_QD_disciplina_agenda(self, df_mapping, sheet_QD, error_dict):
        print("start checking if foreach agenda there is the same Disciplina for all the QD")
        #tutti i QD di un agenda hanno la stessa disciplina
        error_dict.update({
            'error_QD_disciplina_agenda': [],
            'error_QD_descrizione_disciplina_agenda': [],
            'error_disciplina_mancante' : [],
            'error_discipline_agende_diverse': []
        }) 

        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel

        QD_disci_dict_error = {}
        QD_descri_disci_dict_error = {}
        agenda_disci_dict_error = {}
       
        #file_contents=self.file_data.getvalue()
        #for row in range(sheet_mapping.nrows):
        #print(sheet_mapping.row_values(row))
        wb = xlrd.open_workbook(self.file_data)
        sheet_mapping = wb.sheet_by_index(self.work_index_sheet)
        #disciplina_QD_column = sheet_QD[['Cod Disciplina','Codice Quesito']]
        #print("disciplina_QD_column: %s", disciplina_QD_column)
        agende_viewed = []
        for index, row in df_mapping.iterrows():
            disci_flag_QD = False
            descri_disci_flag_QD = False
            disci_flag_agenda = False
            if row[self.work_codice_agenda_siss] not in agende_viewed and row[self.work_codice_QD] is not None:
                searchedAgenda = row[self.work_codice_agenda_siss]
                #disciplina_mapping_row = row[self.work_codici_disciplina_catalogo]
                #descrizione_disciplina_mapping_row = row[self.work_descrizione_disciplina_catalogo]
                #prendo tutte le agende con lo stesso codice
                result = self.findCell_agenda(sheet_mapping, searchedAgenda, self.work_index_codice_SISS_agenda) #prendo tutte le righe con questa agenda
                print("start iterate each excel row")
                if result != -1:
                    result_disciplina_last = ""
                    agende_error_list = []
                    descrizione_discipline_error_list = []
                    discipline_error_list = []
                    for res in result: #per ogni risultato controllo che ci sia la stessa disciplina
                        r = res.split("|")[0] #row agenda
                        c = res.split("|")[1] #column agenda
                        #print("start iterate result sheet_mapping")
                        result_QD = sheet_mapping.cell(int(r), self.work_index_codice_QD).value.split(self.work_delimiter) #QD
                        if result_QD is not None:
                            for QD in result_QD:
                                #print("start iterate QD")
                                QD = QD.strip()
                                if QD != "":
                                    short_sheet = sheet_QD.loc[sheet_QD["Codice Quesito"] == QD]
                                    di_list = []
                                    for ss in short_sheet["Cod Disciplina"].values: #risolvo problema discipline con codici multipli
                                        #print("start iterate cod disciplina" + ss)
                                        if "\n" in ss:
                                            ss = str(ss)
                                            di_list = di_list + ss.split("\n")
                                        else:
                                            di_list.append(ss)
                                    disciplina_mapping_row = df_mapping[self.work_codici_disciplina_catalogo].iloc[int(r)-1]
                                    if disciplina_mapping_row != "":
                                        descrizione_disciplina_mapping_row = df_mapping[self.work_descrizione_disciplina_catalogo].iloc[int(r)-1]
                                        if disciplina_mapping_row not in di_list:
                                            disci_flag_QD = True
                                            if str(int(r)+1) not in agende_error_list:
                                                agende_error_list.append(str(int(r)+1))
                                                QD_disci_dict_error[str(int(r)+1)] = "QD: "+ QD + " non appartiene alla disciplina: " + disciplina_mapping_row
                                                print("disciplina_mapping_row: " + disciplina_mapping_row + ", QD not in disciplina: " + QD)
                                        if descrizione_disciplina_mapping_row not in short_sheet["Descrizione disciplina"].values:
                                            descri_disci_flag_QD = True
                                            if str(int(r)+1) not in descrizione_discipline_error_list:
                                                descrizione_discipline_error_list.append(str(int(r)+1))
                                                QD_descri_disci_dict_error[str(int(r)+1)] = "La descrizione disciplina: " + descrizione_disciplina_mapping_row + "non è presente sul catalogo SISS"
                                    else:
                                        if str(int(r)-1) not in error_dict['error_disciplina_mancante']:
                                            error_dict['error_disciplina_mancante'].append(str(int(r)-1)) 

                        result_disciplina = sheet_mapping.cell(int(r), self.work_index_codici_disciplina_catalogo).value #disciplina da catalogo
                        if result_disciplina != "":
                            if result_disciplina_last != "": #se non è la prima iterazione
                                if result_disciplina != result_disciplina_last:
                                    disci_flag_agenda = True
                                    discipline_error_list.append(str(int(r)+1))
                                    agenda_disci_dict_error = self.update_list_in_dict(agenda_disci_dict_error, str(int(r)+1), "agenda con discipline diverse: "+ result_disciplina + " e " + result_disciplina_last)
                                    #print("result_disciplina: " + result_disciplina + ", result_disciplina_last: " + result_disciplina_last)
                                else: 
                                    result_disciplina_last = result_disciplina    
                
                    if disci_flag_QD == True: #se la disciplina non è rilevata nel catalogo allora è errore
                        for age in agende_error_list:
                            error_dict['error_QD_disciplina_agenda'].append(age)
                            
                    if disci_flag_agenda == True: #se la disciplina è diversa in una stessa agenda, allora è errore
                        for age in discipline_error_list:
                            error_dict['error_discipline_agende_diverse'].append(age)
                else:
                    error_dict['error_disciplina_mancante'].append(str(int(index)+2))     #inserisco la riga senza disciplina negli errori

                agende_viewed.append(searchedAgenda)

        out1 = ""
        out_message = ""
        for ind in error_dict['error_QD_disciplina_agenda']:
            out1 = out1 + "at index: " + ind + ", error: " + QD_disci_dict_error[ind] + ", \n"
            out_message = "__> {}".format(QD_disci_dict_error[ind])
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        
        out2 = ""
        for ind in error_dict['error_discipline_agende_diverse']:
            out2 = out2 + "at index: " + ind + ", error: " + ", ".join(agenda_disci_dict_error[ind]) + ", \n"
            out_message = "__> {}".format(", ".join(agenda_disci_dict_error[ind]))
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        
        out3 = ""
        for ind in error_dict['error_disciplina_mancante']:
            out3 = out3 + "at index: " + ind + "disciplina mancante, \n"
            out_message = "__> Disciplina mancante per la seguente prestazione"
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        
        out4 = ""
        for ind in error_dict['error_QD_descrizione_disciplina_agenda']:
            out4 = out4 + "at index: " + ind + ", error: " + QD_descri_disci_dict_error[ind] + ", \n"
            out_message = "__> {}".format(QD_descri_disci_dict_error[ind])
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        
    
        self.output_message = self.output_message + "\nerror_QD_disciplina_agenda: \n" + out1
        self.output_message = self.output_message + "\nerror_discipline_agende_diverse: \n" + out2
        self.output_message = self.output_message + "\nerror_disciplina_mancante: \n" + out3
        self.output_message = self.output_message + "\error_QD_descrizione_disciplina_agenda: \n" + out4

        xfile.save(self.file_data) 
        return error_dict

    def ck_QD_sintassi(self, df_mapping, error_dict):
        print("start checking if there is ',' separator between each QD defined")
        error_dict.update({
            'error_QD_caratteri_non_consentiti': [],
            'error_QD_spazio_bordi': [],
            'error_QD_spazio_internamente': [],
        })
        string_check = re.compile('1234567890,Q;') #lista caratteri ammessi

        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel

        for index, row in df_mapping.iterrows():
            #print("QD: " + row["Codice Quesito Diagnostico"])
            if row[self.work_abilitazione_esposizione_siss] != None: 
                if row[self.work_abilitazione_esposizione_siss] == "S": 
                    #flag_error = False
                    if row[self.work_codice_QD] is not None:
                        row_replace = row[self.work_codice_QD].replace(" ", "")
                        if " " in row[self.work_codice_QD]:
                            if " " in row_replace.strip():
                                #print("string contain space inside the string")
                                error_dict['error_QD_spazio_internamente'].append(str(int(index)+2))
                            else:
                                #print("string contain space in the border")
                                error_dict['error_QD_spazio_bordi'].append(str(int(index)+2))
                        elif(string_check.search(row_replace) != None):
                            #print("String contains other Characters.")
                            error_dict['error_QD_caratteri_non_consentiti'].append(str(int(index)+2))
                            #flag_error = True
                       
        out_message = ""
        for ind in error_dict['error_QD_caratteri_non_consentiti']:
            out_message = "__> QD presentano errori di sintassi: rilevati caratteri non consentiti"
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        for ind in error_dict['error_QD_spazio_bordi']:
            out_message = "__> QD presentano errori di sintassi: rilevati degli spazi alle estremità del contenuto della cella"
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        for ind in error_dict['error_QD_spazio_internamente']:
            out_message = "__> QD presentano errori di sintassi: rilevati degli spazi all'interno del contenuto della cella"
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message

        xfile.save(self.file_data) 
        return error_dict

    def ck_QD_descrizione(self, df_mapping, sheet_QD, error_dict):
        print("start checking if there are the relative QD description")
        error_dict.update({
            'error_QD_descrizione': [],
            'error_QD_descrizione_space_bordo': [],
            'error_QD_descrizione_space_interno': []
        })

        QD_dict_error_1 = {}
        QD_dict_error_2 = {}
        QD_dict_error_3 = {}
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel

        for index, row in df_mapping.iterrows():
            if row[self.work_abilitazione_esposizione_siss] == "S":
                QD_string = row[self.work_codice_QD].split(self.work_delimiter)
                description_list = row[self.work_descrizione_QD]#.split(self.work_delimiter)
                flag_error = False
                '''if len(QD_string) != len(row[self.work_descrizione_QD].split(self.work_delimiter)):
                    print("il numero di descrizioni è diverso dal numero di QD all'indice " + str(index))
                    flag_error = True
                    QD_dict_error_3 = self.update_list_in_dict(QD_dict_error_3, str(int(index)+2), row[self.work_codice_QD])
                '''  

                if QD_string is not None:
                    for QD in QD_string:
                        QD = QD.strip()
                        if QD != "":
                            QD_catalogo = sheet_QD.loc[sheet_QD["Codice Quesito"] == QD]  
                            #print("QD: " + str(QD)) 
                                                        
                            if description_list != description_list.strip(): #there is a space in the beginning or in the end
                                error_dict['error_QD_descrizione_space_bordo'].append(str(int(index)+2))
                                logging.error("ERROR SPACE BORDI: controllare QD: " + QD + " all'indice: " + str(int(index)+2))
                                QD_dict_error_1 = self.update_list_in_dict(QD_dict_error_1, str(int(index)+2), QD)
                                description_list = description_list.strip()

                            if (" "+self.work_delimiter) in description_list or (self.work_delimiter+" ") in description_list:
                                #print("print QD_catalogo2:" + QD_catalogo)
                                #print("controllare manualmente qual'è il problema")
                                print("QD: " + QD + ", Quesiti Diagnostici size:" + str(QD_catalogo.size) + ", description_list: %s", description_list)
                                logging.error("ERROR SPACE INTERNO: controllare QD: " + QD + " all'indice: " + str(int(index)+2))
                                error_dict['error_QD_descrizione_space_interno'].append(str(int(index)+2))
                                QD_dict_error_2 = self.update_list_in_dict(QD_dict_error_2, str(int(index)+2), QD)
                                description_list = description_list.replace(self.work_delimiter+" ", self.work_delimiter) #elimino spazio dopo  del separatore
                                description_list = description_list.replace(" "+self.work_delimiter, self.work_delimiter) #elimino spazio prima del separatore
                            try:
                                if QD_catalogo["Quesiti Diagnostici"].values[0] not in description_list.split(self.work_delimiter):
                                    print("la descrizione QD non è presente all'indice " + str(int(index)+2))
                                    #print("QD: " + QD + ", Quesiti Diagnostici: " + QD_catalogo["Quesiti Diagnostici"].values[0] + ", Description_list: %s", description_list)
                                    logging.error("ERROR DESCRIZIONE: controllare QD: " + QD + " all'indice: " + str(int(index)+2))
                                    flag_error = True
                                    QD_dict_error_3 = self.update_list_in_dict(QD_dict_error_3, str(int(index)+2), QD)
                            except: #togliere try/catch e gestire gli spazi nell'if sopra
                                #print("print QD_catalogo2:" + QD_catalogo)
                                flag_error = True
                                QD_dict_error_3 = self.update_list_in_dict(QD_dict_error_3, str(int(index)+2), QD)
                                print("controllare manualmente qual'è il problema all'indice: " + str(int(index)+2))
                                #print("QD: " + QD + ", Quesiti Diagnostici size:" + str(QD_catalogo.size) + ", Description_list: %s", description_list)
                                #logging.error("controllare manualmente il QD: " + QD + " all'indice: " + str(int(index)+2))
                                #error_dict['error_QD_descrizione_space_interno'].append(str(int(index)+2))  
                                      
                            
                if flag_error == True:
                   error_dict['error_QD_descrizione'].append(str(int(index)+2))  

        out_message = ""
        for ind in error_dict['error_QD_descrizione']:
            out_message = "__> Descrizione dei QD: '{}' non trovati nel catalogo SISS".format(", ".join(QD_dict_error_3[ind]))
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        for ind in error_dict['error_QD_descrizione_space_bordo']:
            out_message = "__> Descrizione dei QD: '{}' presentano spazi alle estremità della cella ".format(", ".join(QD_dict_error_1[ind]))
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        for ind in error_dict['error_QD_descrizione_space_interno']:
            out_message = "__> Descrizione dei QD: '{}' presentano spazi non consentiti tra i QD specificati".format(", ".join(QD_dict_error_2[ind]))
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message

        xfile.save(self.file_data) 
        return error_dict

    def ck_QD_operatori_logici(self, df_mapping, error_dict):
        print("start checking if there are the same logic op. for each agenda")
        error_dict.update({
                        'error_QD_operatori_logici': [],
                        'error_QD_operatori_logici_mancante': []})
        QD_dict_error = {}

        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel

        wb = xlrd.open_workbook(self.file_data)
        sheet_mapping = wb.sheet_by_index(self.work_index_sheet)
        print("sheet caricato")

        #agenda = df_mapping[self.work_codice_agenda_siss].iloc[2]
        #last_OP = df_mapping[self.work_operatore_logico_QD].iloc[2]

        agende_checked = []
        
        for index, row in df_mapping.iterrows():
            if row[self.work_abilitazione_esposizione_siss] == "S" and row[self.work_codice_QD].strip() != "":
                if row[self.work_operatore_logico_QD] is not "":
                    agenda_SISS = str(row[self.work_codice_agenda_siss])
                    agenda_SISS = agenda_SISS.strip()
                    if agenda_SISS not in agende_checked:
                        result = self.findCell(sheet_mapping, agenda_SISS, self.work_index_codice_SISS_agenda)

                        if result != -1:
                            for res in result:
                                r = res.split("#")[0]
                                c = res.split("#")[1]
                                resultOP = sheet_mapping.cell(int(r), self.work_index_operatore_logico_QD).value
                                if row[self.work_operatore_logico_QD] != resultOP and (row[self.work_operatore_logico_QD] != "" or resultOP != ""):
                                    QD_dict_error = self.update_list_in_dict(QD_dict_error, str(int(r)+1), agenda_SISS)
                                    error_dict['error_QD_operatori_logici'].append(str(int(r)+1))
                                    print("error QD OP at index:" +  str(int(r)+1))
                    agende_checked.append(agenda_SISS)

                elif row[self.work_operatore_logico_distretto] is "" and row[self.work_codice_distretto] is not "": 
                    error_dict['error_QD_operatori_logici_mancante'].append(str(int(index)+2))

                '''if row[self.work_codice_agenda_siss] == agenda:
                    if row[self.work_operatore_logico_QD] == "":
                        error_dict['error_QD_operatori_logici'].append(str(int(index)+2))
                    else:  
                        if row[self.work_operatore_logico_QD] != last_OP:
                            print("error OP at index:" +  str(int(index)+2))
                            error_dict['error_QD_operatori_logici'].append(str(int(index)+2))
                else: 
                    agenda = row[self.work_codice_agenda_siss]
                    last_OP = row[self.work_operatore_logico_QD]'''

        out_message = ""
        for ind in error_dict['error_QD_operatori_logici']:
            out_message = "__> QD: trovato errore sull'operatore logico. Controllare se è presente e che è conforme nell'agenda"
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message
        for ind in error_dict['error_QD_operatori_logici_mancante']:
            out_message = "__> QD: Operatore logico assente"
            if sheet[self.work_alert_column+ind].value is not None:
                sheet[self.work_alert_column+ind] = str(sheet[self.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.work_alert_column+ind] = out_message

        xfile.save(self.file_data)
        return error_dict