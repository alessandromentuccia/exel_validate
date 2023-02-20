import os
from distutils.log import error
import logging
from collections import OrderedDict
from functools import reduce
from pathlib import Path
from typing import Dict, List
import pandas as pd
import numpy as np
import xlrd
import openpyxl
from xlsxwriter.utility import xl_rowcol_to_cell
import matplotlib.pyplot as plt
from dotenv import load_dotenv

from flaskr.check.check_QD import Check_QD
from flaskr.check.check_metodiche import Check_metodiche 
from flaskr.check.check_distretti import Check_distretti
from flaskr.check.check_priorita import Check_priorita
from flaskr.check.check_prestazione import Check_prestazione
from flaskr.check.check_canali import Check_canali
from flaskr.check.check_agende_interne import Check_agende_interne
from flaskr.check.check_inviante import Check_inviante
from flaskr.report_creation.rep_creation import Report_Creation #rep_creation.

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

RESULT_VALIDATION = "..\check_excel_result.txt"
load_dotenv()
CAT_NAME = os.getenv("CAT_SISS")
ROOT_DIR = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))

class Check_action():

    file_name = ""
    file_data = {}
    controls_setted = {}
    catalogo = OrderedDict()
    flag_check_list = []
    error_generate_list = {
        1 : "Lo Sheet indicato non è presente nel file .xlsx: ricontrollare il file di configurazione ed il file excel"
    }
    validation_results = ""
    output_message = ""
    
    #work_sheet = "" #sheet di lavoro di df_mapping
    #work_codice_prestazione_siss = ""
    #work_descrizione_prestazione_siss = ""
    #work_codice_agenda_siss = ""
    #work_casi_1_n = ""
    #work_abilitazione_esposizione_siss = ""
    #work_prenotabile_siss = ""
    #work_codici_disciplina_catalogo = ""
    #work_descrizione_disciplina_catalogo = ""
    #work_codice_QD = ""
    #work_descrizione_QD = ""
    #work_operatore_logico_QD = ""
    #work_codice_metodica = ""
    #work_descrizione_metodica = ""
    #work_codice_distretto = ""
    #work_descrizione_distretto = ""
    #work_operatore_logico_distretto = ""
    #work_priorita_U = ""
    #work_priorita_primo_accesso_D = ""
    #work_priorita_primo_accesso_P = ""
    #work_priorita_primo_accesso_B = ""
    #work_accesso_programmabile_ZP = ""
    #work_combinata = ""
    #work_codice_agenda_interno = ""
    #work_codice_prestazione_interno = ""
    #work_inviante = ""
    #work_accesso_farmacia = ""
    #work_accesso_CCR = ""
    #work_accesso_cittadino = ""
    #work_accesso_MMG = ""
    #work_accesso_amministrativo = ""
    #work_accesso_PAI = ""
    #work_gg_preparazione = ""
    #work_gg_refertazione = ""
    #work_nota_operatore = ""

    #work_index_codice_QD = 0
    #work_index_codice_SISS_agenda = 0
    #work_index_abilitazione_esposizione_SISS = 0
    #work_index_codice_prestazione_SISS = 0
    #work_index_operatore_logico_distretto = 0
    #work_index_codici_disciplina_catalogo = 0
    #work_index_operatore_logico_QD = 0

    #work_alert_column = ""
    #work_delimiter = ""


    def __init__(self, data, excel_file):
        #self.output_message = ""
        #with open("./flaskr/config_validator.yml", "rt", encoding='utf8') as yamlfile:
        #    data = yaml.load(yamlfile, Loader=yaml.FullLoader)
        #logger.debug(data)
        self.work_sheet = data[0]["work_column"]["work_sheet"] 
        self.work_N1 = data[0]["work_column"]["work_N1"] 
        self.work_N2 = data[0]["work_column"]["work_N2"]
        self.work_descrizione_N1 = data[0]["work_column"]["work_descrizione_N1"]
        self.work_descrizione_N2 = data[0]["work_column"]["work_descrizione_N2"]
        self.work_codice_prestazione_siss = data[0]["work_column"]["work_codice_prestazione_siss"]
        self.work_descrizione_prestazione_siss = data[0]["work_column"]["work_descrizione_prestazione_siss"]
        self.work_codice_agenda_siss = data[0]["work_column"]["work_codice_agenda_siss"]
        self.work_casi_1_n = data[0]["work_column"]["work_casi_1_n"]
        self.work_abilitazione_esposizione_siss = data[0]["work_column"]["work_abilitazione_esposizione_siss"]
        self.work_prenotabile_siss = data[0]["work_column"]["work_prenotabile_siss"]
        self.work_codici_disciplina_catalogo = data[0]["work_column"]["work_codici_disciplina_catalogo"]
        self.work_descrizione_disciplina_catalogo = data[0]["work_column"]["work_descrizione_disciplina_catalogo"]
        self.work_codice_QD = data[0]["work_column"]["work_codice_QD"]
        self.work_descrizione_QD = data[0]["work_column"]["work_descrizione_QD"]
        self.work_operatore_logico_QD = data[0]["work_column"]["work_operatore_logico_QD"]
        self.work_codice_metodica = data[0]["work_column"]["work_codice_metodica"]
        self.work_descrizione_metodica = data[0]["work_column"]["work_descrizione_metodica"]
        self.work_codice_distretto = data[0]["work_column"]["work_codice_distretto"]
        self.work_descrizione_distretto = data[0]["work_column"]["work_descrizione_distretto"]
        self.work_operatore_logico_distretto = data[0]["work_column"]["work_operatore_logico_distretto"]
        self.work_priorita_U = data[0]["work_column"]["work_priorita_U"]
        self.work_priorita_primo_accesso_D = data[0]["work_column"]["work_priorita_primo_accesso_D"]
        self.work_priorita_primo_accesso_P = data[0]["work_column"]["work_priorita_primo_accesso_P"]
        self.work_priorita_primo_accesso_B = data[0]["work_column"]["work_priorita_primo_accesso_B"]
        self.work_accesso_programmabile_ZP = data[0]["work_column"]["work_accesso_programmabile_ZP"]
        self.work_combinata = data[0]["work_column"]["work_combinata"]
        self.work_codice_agenda_interno = data[0]["work_column"]["work_codice_agenda_interno"]
        self.work_codice_prestazione_interno = data[0]["work_column"]["work_codice_prestazione_interno"]
        self.work_inviante = data[0]["work_column"]["work_inviante"]
        self.work_accesso_farmacia = data[0]["work_column"]["work_accesso_farmacia"]
        self.work_accesso_CCR = data[0]["work_column"]["work_accesso_CCR"]
        self.work_accesso_cittadino = data[0]["work_column"]["work_accesso_cittadino"]
        self.work_accesso_MMG = data[0]["work_column"]["work_accesso_MMG"]
        self.work_accesso_amministrativo = data[0]["work_column"]["work_accesso_amministrativo"]
        self.work_accesso_PAI = data[0]["work_column"]["work_accesso_PAI"]
        self.work_gg_preparazione = data[0]["work_column"]["work_gg_preparazione"]
        self.work_gg_refertazione = data[0]["work_column"]["work_gg_refertazione"]
        self.work_nota_operatore = data[0]["work_column"]["work_nota_operatore"]
        self.work_nota_agenda = data[0]["work_column"]["work_nota_agenda"]
        self.work_nota_revoca = data[0]["work_column"]["work_nota_revoca"]
        self.work_disciplina = data[0]["work_column"]["work_disciplina"]
        self.work_sesso = data[0]["work_column"]["work_sesso"]
        self.work_eta_min = data[0]["work_column"]["work_eta_min"]
        self.work_eta_max = data[0]["work_column"]["work_eta_max"]

        self.work_alert_column = data[1]["work_index"]["work_alert_column"]
        try:
            self.work_delimiter = data[2]["work_separator"]["work_delimiter"]
        except:
            self.work_delimiter = "," #valore di default
        self.file_data = excel_file


    def initializer(self, checked_dict):

        print("FASE 0: precheck")
        self.check_sheet_existance()
        if self.validation_results != "":
            return self.validation_results 
        #pd.set_option("display.max_rows", None, "display.max_columns", None)
        df_mapping = pd.read_excel(self.file_data, sheet_name=self.work_sheet, converters={self.work_codici_disciplina_catalogo: str, self.work_codice_prestazione_siss: str, self.work_codice_metodica: str, self.work_codice_distretto: str}).replace(np.nan, '', regex=True)
        #print ("print JSON")
        #print(sh)
        print(checked_dict)
        self.controls_setted = checked_dict

        #catalogo_dir = os.path.join(ROOT_DIR, 'CCR-BO-CATGP#01_Codifiche attributi catalogo GP++_110322.xlsx')
        catalogo_dir = os.path.join(ROOT_DIR, CAT_NAME) #catalogo SISS
        
        sheet_QD = pd.read_excel(catalogo_dir, sheet_name='QD', converters={"Cod Disciplina": str})
        sheet_Metodiche = pd.read_excel(catalogo_dir, sheet_name='METODICHE', converters={"Codice SISS": str, "Codice Metodica": str})
        sheet_Distretti = pd.read_excel(catalogo_dir, sheet_name='DISTRETTI', converters={"Codice SISS": str, "Codice Distretto": str})
        
        print("sheet_QD caricato\n")
        #print(sheet_QD)
        print("sheet_Metodiche caricato\n")
        #print(sheet_Metodiche)
        print("sheet_Distretti caricato\n")
        #print(sheet_Distretti)

        self.analizer(df_mapping, sheet_QD, sheet_Metodiche, sheet_Distretti)

        return self.validation_results 


    def analizer(self, df_mapping, sheet_QD, sheet_Metodiche, sheet_Distretti):

        print('Start analisys:\n', df_mapping)
        QD_error = {}
        if self.controls_setted["Quesiti"] == 1:
            print("Fase 1: Controllo Quesiti selezionato") #FASE 1: CONTROLLO I QUESITI DIAGNOSTICI
            QD_error = self.check_qd(df_mapping, sheet_QD)
        metodiche_error = {}
        if self.controls_setted["Metodiche"] == 1:
            print("Fase 2: Controllo Metodiche selezionato") #FASE 2: CONTROLLO LE METODICHE
            metodiche_error = self.check_metodiche(df_mapping, sheet_Metodiche)
        distretti_error = {}
        if self.controls_setted["Distretti"] == 1:    
            print("Fase 3: Controllo Distretti selezionato") #FASE 3: CONTROLLO I DISTRETTI
            distretti_error = self.check_distretti(df_mapping, sheet_Distretti)
        priorita_error = {}
        if self.controls_setted["Priorita"] == 1:
            print("Fase 4: Controllo Priorita selezionato") #FASE 4: CONTROLLO LE PRIORITA'
            priorita_error = self.check_priorita(df_mapping)
        prestazione_error = {}
        if self.controls_setted["Prestazione"] == 1:
            print("Fase 5: Controllo Prestazione selezionato") #FASE 5: CONTROLLO UNIVOCITA' PRESTAZIONI'
            prestazione_error = self.check_prestazione(df_mapping)
        canali_error = {}
        if self.controls_setted["Canali"] == 1:
            print("Fase 6: Controllo Canali selezionato")
            canali_error = self.check_canali(df_mapping)
        inviante_error = {}
        if self.controls_setted["Inviante"] == 1:
            print("Fase 7: Controllo Inviante selezionato")
            inviante_error = self.check_inviante(df_mapping)
        print("Fase Vale Validator")
        
        '''#catalogo_dir = os.path.join(ROOT_DIR, 'CCR-BO-CATGP#01_Codifiche attributi catalogo GP++_110322.xls')
        catalogo_dir = os.path.join(ROOT_DIR, CAT_NAME)
        wb = xlrd.open_workbook(catalogo_dir)
        sheet_QD_OW = wb.sheet_by_index(1)
        sheet_Metodiche_OW = wb.sheet_by_index(2)
        sheet_Distretti_OW = wb.sheet_by_index(3)
        QD_validator_error = {}
        metodiche_validator_error = {}
        distretti_validator_error = {}
        #QD_validator_error = Validator_v.ck_QD_description(self, df_mapping, sheet_QD_OW)
        #metodiche_validator_error = Validator_v.ck_metodiche_description(self, df_mapping, sheet_Metodiche_OW)
        #distretti_validator_error = Validator_v.ck_distretti_description(self, df_mapping, sheet_Distretti_OW)'''


        error_dict = {
            "QD_error": QD_error,
            "metodiche_error": metodiche_error,
            "distretti_error": distretti_error,
            "priorita_error": priorita_error,    
            "prestazione_error": prestazione_error,
            '''"QD_validator_error": QD_validator_error,
            "metodiche_validator_error": metodiche_validator_error,
            "distretti_validator_error": distretti_validator_error,'''
            "canali_error": canali_error,
            "inviante_error": inviante_error
        }

        self._validation(error_dict, df_mapping)

        

    def check_sheet_existance(self):
        print("check the used column name of the excel file")
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        if self.work_sheet in xfile.sheetnames:
            print('sheet exists')
            self.validation_results = ""
        else: 
            self.validation_results = self.error_generate_list[1]


    def check_qd(self, df_mapping, sheet_QD):
        print("start checking QD") #Codice Quesito Diagnostico
        #controllo se per ogni Agenda sono inseriti gli stessi QD
        error_dict = {}

        error_QD_sintassi = Check_QD.ck_QD_sintassi(self, df_mapping, error_dict)
        error_QD_agenda = Check_QD.ck_QD_agenda(self, df_mapping, error_QD_sintassi)
        error_QD_disciplina_agenda = Check_QD.ck_QD_disciplina_agenda(self, df_mapping, sheet_QD, error_QD_agenda)
        error_QD_descrizione = Check_QD.ck_QD_descrizione(self, df_mapping, sheet_QD, error_QD_disciplina_agenda)
        error_QD_operatori_logici = Check_QD.ck_QD_operatori_logici(self, df_mapping, error_QD_descrizione)

        error_dict = error_QD_operatori_logici

        return error_dict

    def check_metodiche(self, df_mapping, sheet_Metodiche):
        print("start checking Metodiche")
        error_dict = {}

        self.output_message = self.output_message + "\nErrori presenti nelle metodiche e riportate attraverso gli indici:\n"

        error_metodica_sintassi = Check_metodiche.ck_metodica_sintassi(self, df_mapping, error_dict)
        error_metodica_inprestazione = Check_metodiche.ck_metodica_inprestazione(self, df_mapping, sheet_Metodiche, error_metodica_sintassi)
        #error_metodica_descrizione = Check_metodiche.ck_metodica_descrizione(self, df_mapping, sheet_Metodiche, error_metodica_inprestazione)

        error_dict = error_metodica_inprestazione

        #print("error_dict: %s", error_dict)
        return error_dict

    def check_distretti(self, df_mapping, sheet_Distretti):
        print("start checking Distretti")
        error_dict = {}

        error_distretti_sintassi = Check_distretti.ck_distretti_sintassi(self, df_mapping, error_dict)
        error_distretti_inprestazione = Check_distretti.ck_distretti_inprestazione(self, df_mapping, sheet_Distretti, error_distretti_sintassi)
        #error_distretti_descrizione = Check_distretti.ck_distretti_descrizione(self, df_mapping, sheet_Distretti, error_distretti_inprestazione)
        error_distretti_operatori_logici = Check_distretti.ck_distretti_operatori_logici(self, df_mapping, error_distretti_inprestazione)

        error_dict = error_distretti_operatori_logici

        return error_dict

    def check_priorita(self, df_mapping):
        print("start checking priorità e tipologie di accesso")
        error_dict = {}

        error_prime_visite = Check_priorita.ck_prime_visite(self, df_mapping, error_dict)
        error_controlli = Check_priorita.ck_controlli(self, df_mapping, error_prime_visite)
        error_esami_strumentali =Check_priorita.ck_esami_strumentali(self, df_mapping, error_controlli)

        error_dict = error_esami_strumentali

        return error_dict

    def check_prestazione(self, df_mapping):
        print("start checking univocità delle prestazioni")
        error_dict = {}

        error_casi_1N = Check_prestazione.ck_casi_1n(self, df_mapping, error_dict)
        error_prestazione = Check_prestazione.ck_prestazione(self, df_mapping, error_casi_1N)
        error_prestazione_non_prenotabile = Check_prestazione.ck_prestazione_nonprenotabile(self, df_mapping, error_prestazione)

        error_dict = error_prestazione_non_prenotabile

        return error_dict

    def check_canali(self, df_mapping):
        print("start checking canali di accesso")

        error_dict = {}
        error_canali_vuoti = Check_canali.ck_canali_vuoti(self, df_mapping, error_dict)
        error_canali_PAI = Check_canali.ck_canali_PAI(self, df_mapping, error_canali_vuoti)
        error_canali_abilitati = Check_canali.ck_canali_abilitati(self, df_mapping, error_canali_PAI)

        error_dict = error_canali_abilitati

        return error_dict

    def check_inviante(self, df_mapping):
        print("start checking inviante")

        error_dict = {}

        error_inviante = Check_inviante.ck_inviante(self, df_mapping, error_dict)

        error_dict = error_inviante

        return error_dict

    def _validation(self, error_dict, df_mapping):
        print("questi sono gli errori indivuduati e separati per categoria:\n %s", error_dict)
        #self.output_message = self.output_message + "\n" + json.dumps(error_dict)
        '''df = pd.DataFrame(rows_list)
        with pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name='new_mapping', index=False)'''
        print("\nPer osservare i risultati ottenuti, controllare il file prodotto: check_excel_result.txt")
        file = open(RESULT_VALIDATION, "w") 
        file.write(self.output_message)
        file.close() 
        try:
            #Report_creation = Report_Creation()
            Report_creation = Report_Creation(  df_mapping, 
                                                self.file_data,
                                                self.work_sheet,
                                                self.work_N1,
                                                self.work_N2,
                                                self.work_descrizione_N1,
                                                self.work_descrizione_N2,
                                                self.work_codice_prestazione_siss,
                                                self.work_descrizione_prestazione_siss,
                                                self.work_codice_agenda_siss,
                                                self.work_casi_1_n,
                                                self.work_abilitazione_esposizione_siss,
                                                self.work_prenotabile_siss,
                                                self.work_codici_disciplina_catalogo,
                                                self.work_descrizione_disciplina_catalogo,
                                                self.work_codice_QD,
                                                self.work_codice_metodica,
                                                self.work_codice_distretto,
                                                self.work_priorita_U,
                                                self.work_priorita_primo_accesso_D,
                                                self.work_priorita_primo_accesso_P,
                                                self.work_priorita_primo_accesso_B,
                                                self.work_accesso_programmabile_ZP,
                                                self.work_combinata,
                                                self.work_codice_agenda_interno,
                                                self.work_codice_prestazione_interno,
                                                self.work_inviante,
                                                self.work_accesso_farmacia,
                                                self.work_accesso_CCR,
                                                self.work_accesso_cittadino,
                                                self.work_accesso_MMG,
                                                self.work_accesso_amministrativo,
                                                self.work_accesso_PAI,
                                                self.work_gg_preparazione,
                                                self.work_gg_refertazione,
                                                self.work_nota_operatore,
                                                self.work_nota_agenda,
                                                self.work_nota_revoca,
                                                self.work_disciplina,
                                                self.work_sesso,
                                                self.work_eta_min,
                                                self.work_eta_max,
                                                self.work_alert_column,
                                                self.work_delimiter,
                                                error_dict)
            Report_creation.get_report()

        except Exception as e:
            logger.error('Failed to upload to ftp: '+ str(e))
            print("Errore in Report Creation")
        '''try:
            xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
            #sheet_mapping = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel mapping
            try:
                sheet_report = xfile.get_sheet_by_name('Report Validazione') #recupero sheet excel report validazione 
            except:
                #creo Report Validazione se non esiste
                sheet_report = xfile.create_sheet('Report Validazione')
            sheet_report["A1"] = "Report Validazione"

            xfile.save(self.file_data) 
        except:
            print("non esiste file o sheet")'''
        
        

    def findCell(self, sh, searchedValue, start_col):
        result_coord = []
        result_value = []

        for row in range(sh.nrows):
            for col in range(start_col, start_col+1):
                myCell = sh.cell(row, col)
                myValue = sh.cell(row, self.work_index_codice_prestazione_SISS)
                abilita = sh.cell(row, self.work_index_abilitazione_esposizione_SISS)
                if myCell.value == searchedValue and abilita.value == "S":
                    result_coord.append(str(row) + "#" + str(col))
                    result_value.append(myValue.value)
                    #return row, col#xl_rowcol_to_cell(row, col)

        if result_coord == []:
            return -1
        return result_coord#, result_value

    def findCell_agenda(self, sh, searchedValue, start_col):
        result_coord = []
        result_value = []

        for row in range(sh.nrows):
            for col in range(start_col, start_col+1):
                myCell = sh.cell(row, col)
                myValue = sh.cell(row, self.work_index_codice_SISS_agenda) #Codice SISS agenda 15
                #abilita = sh.cell(row, self.work_index_abilitazione_esposizione_SISS-1) #abilitazione esposizione SISS 28
                if myCell.value == searchedValue: # and abilita.value == "S":
                    result_coord.append(str(row) + "|" + str(col))
                    result_value.append(myValue.value)
                    #return row, col#xl_rowcol_to_cell(row, col)

        if result_coord == []:
            return -1
        return result_coord#, result_value
                                    
    def findCell_agenda_II(self, df, searchedValue, name_column):#sheet_mapping, searchedAgenda, self.work_index_codice_SISS_agenda
        result_coord = []

        for rowIndex, row in df.iterrows(): #iterate over rows
            myCell = row[name_column]
            abilita = row[self.work_abilitazione_esposizione_siss]
            if myCell == searchedValue and abilita == "S":
                result_coord.append(str(rowIndex))

        if result_coord == []:
            return -1
        return result_coord

    '''Metodo che aggiunge elemento in una lista esistente o crea la lista nel caso
    non fosse presente'''
    def update_list_in_dict(self, dictio, index, element):
        if index in dictio.keys():
            dictio[index].append(element)
        else:
            dictio[index] = [element]
        return dictio

    '''Metodo per controllare il tipo di una colonna excel'''
    def column_validator():
        return ""

    def list_duplicates(self, seq):
        seen = set()
        seen_add = seen.add
        # adds all elements it doesn't know yet to seen and all other to seen_twice
        seen_twice = set( x for x in seq if x in seen or seen_add(x) )
        # turn the set into a list (as requested)
        return list( seen_twice )


    def initializer_check_agende_interne(self):
        df_mapping = pd.read_excel(self.file_data, sheet_name=self.work_sheet, converters={self.work_codici_disciplina_catalogo: str, self.work_codice_prestazione_siss: str}).replace(np.nan, '', regex=True)
        error = self.analizer_agende_interne(df_mapping)
        error_dict = {
            "error_Aagende_interne": error
        }
        self._validation(error_dict)

    def analizer_agende_interne(self, df_mapping):
        Agende_interne_error = Check_agende_interne.ck_agende_interne(self, df_mapping, {})
        
        return Agende_interne_error