import os
import json
#from distutils.log import error
import logging
from collections import OrderedDict
#from functools import reduce
#from pathlib import Path
#from typing import Dict, List
#import pandas as pd
#import numpy as np
#import xlrd
import openpyxl
#from xlsxwriter.utility import xl_rowcol_to_cell
from dotenv import load_dotenv

#from flaskr.check.check_QD import Check_QD
#from flaskr.check.check_metodiche import Check_metodiche
#from flaskr.check.check_distretti import Check_distretti
#from flaskr.check.check_priorita import Check_priorita
#from flaskr.check.check_prestazione import Check_prestazione
#from flaskr.check.check_canali import Check_canali
#from flaskr.check.check_agende_interne import Check_agende_interne
#from flaskr.check.check_inviante import Check_inviante
from flaskr.report_creation.rep_creation import Report_Creation #rep_creation.

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

RESULT_VALIDATION = "..\check_excel_result.txt"
load_dotenv()
CAT_NAME = os.getenv("CAT_SISS")
ROOT_DIR = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))


class Check_action():


    controls_setted = {}
    flag_check_list = []
    error_generate_list = {
        1 : "Lo Sheet indicato non è presente nel file .xlsx: ricontrollare il file di configurazione ed il file excel",
        2 : "Nello Sheet indicato non è presente il campo "
    }
    validation_results = ""
    output_message = ""
    print("ROOT_DIR: " + ROOT_DIR+"\n")
    print("CAT_NAME: " + str(CAT_NAME)+"\n")
    catalogo_dir = os.path.join(ROOT_DIR, CAT_NAME) #catalogo attributi


    def __init__(self, file):#, data, excel_file, checked_dict):
        #self.output_message = ""
        #with open("./flaskr/config_validator.yml", "rt", encoding='utf8') as yamlfile:
        #    data = yaml.load(yamlfile, Loader=yaml.FullLoader)
        #logger.debug(data)

        #ottengo directory catalogo attributi
        #catalogo_dir = os.path.join(ROOT_DIR, 'CCR-BO-CATGP#01_Codifiche attributi catalogo GP++_110322.xlsx')
        self.file = file
        #self.df_mapping = pd.read_excel(self.file_data, sheet_name=self.data.work_sheet, converters={self.data.work_codici_disciplina_catalogo: str, self.data.work_codice_prestazione_siss: str, self.data.work_codice_metodica: str, self.data.work_codice_distretto: str}).replace(np.nan, '', regex=True)
        
        #ottengo i singoli cataloghi QD, metodiche e distretti
        '''self.sheet_QD, self.sheet_Metodiche, self.sheet_Distretti = self.get_catalogo(self.catalogo_dir)

        self.work_sheet = data[0]["work_column"]["work_sheet"]
        self.work_N1 = data[0]["work_column"]["work_N1"]
        self.work_N2 = data[0]["work_column"]["work_N2"]
        self.work_descrizione_N1 = data[0]["work_column"]["work_descrizione_N1"]
        self.work_descrizione_N2 = data[0]["work_column"]["work_descrizione_N2"]
        self.work_codice_prestazione_siss = data[0]["work_column"]["work_codice_prestazione_siss"]
        self.work_descrizione_prestazione_siss = data[0]["work_column"]["work_descrizione_prestazione_siss"]
        self.work_codice_agenda_siss = data[0]["work_column"]["work_codice_agenda_siss"]
        #self.work_casi_1_n = data[0]["work_column"]["work_casi_1_n"]
        self.work_abilitazione_esposizione_siss = data[0]["work_column"]["work_abilitazione_esposizione_siss"]
        self.work_prenotabile_siss = data[0]["work_column"]["work_prenotabile_siss"]
        self.work_codici_disciplina_catalogo = data[0]["work_column"]["work_codici_disciplina_catalogo"]
        self.work_descrizione_disciplina_catalogo = data[0]["work_column"]["work_descrizione_disciplina_catalogo"]
        self.work_codice_QD = data[0]["work_column"]["work_codice_QD"]
        self.work_descrizione_QD = data[0]["work_column"]["work_descrizione_QD"]
        self.work_operatore_logico_QD = data[0]["work_column"]["work_operatore_logico_QD"]
        self.work_codice_metodica = data[0]["work_column"]["work_codice_metodica"]
        self.work_descrizione_metodica = data[0]["work_column"]["work_descrizione_metodica"]
        self.work_codice_distretto = data[0]["work_column"]["work_codice_distretto"]
        self.work_descrizione_distretto = data[0]["work_column"]["work_descrizione_distretto"]
        self.work_operatore_logico_distretto = data[0]["work_column"]["work_operatore_logico_distretto"]
        self.work_priorita_U = data[0]["work_column"]["work_priorita_U"]
        self.work_priorita_primo_accesso_D = data[0]["work_column"]["work_priorita_primo_accesso_D"]
        self.work_priorita_primo_accesso_P = data[0]["work_column"]["work_priorita_primo_accesso_P"]
        self.work_priorita_primo_accesso_B = data[0]["work_column"]["work_priorita_primo_accesso_B"]
        self.work_accesso_programmabile_ZP = data[0]["work_column"]["work_accesso_programmabile_ZP"]
        self.work_combinata = data[0]["work_column"]["work_combinata"]
        self.work_codice_agenda_interno = data[0]["work_column"]["work_codice_agenda_interno"]
        self.work_codice_prestazione_interno = data[0]["work_column"]["work_codice_prestazione_interno"]
        self.work_inviante = data[0]["work_column"]["work_inviante"]
        self.work_accesso_farmacia = data[0]["work_column"]["work_accesso_farmacia"]
        self.work_accesso_CCR = data[0]["work_column"]["work_accesso_CCR"]
        self.work_accesso_cittadino = data[0]["work_column"]["work_accesso_cittadino"]
        self.work_accesso_MMG = data[0]["work_column"]["work_accesso_MMG"]
        self.work_accesso_amministrativo = data[0]["work_column"]["work_accesso_amministrativo"]
        self.work_accesso_PAI = data[0]["work_column"]["work_accesso_PAI"]
        self.work_gg_preparazione = data[0]["work_column"]["work_gg_preparazione"]
        self.work_gg_refertazione = data[0]["work_column"]["work_gg_refertazione"]
        self.work_nota_operatore = data[0]["work_column"]["work_nota_operatore"]
        self.work_nota_agenda = data[0]["work_column"]["work_nota_agenda"]
        self.work_nota_revoca = data[0]["work_column"]["work_nota_revoca"]
        self.work_disciplina = data[0]["work_column"]["work_disciplina"]
        self.work_sesso = data[0]["work_column"]["work_sesso"]
        self.work_eta_min = data[0]["work_column"]["work_eta_min"]
        self.work_eta_max = data[0]["work_column"]["work_eta_max"]

        self.work_alert_column = data[1]["work_index"]["work_alert_column"]
        try:
            self.work_delimiter = data[2]["work_separator"]["work_delimiter"]
        except:
            self.work_delimiter = "," #valore di default
        self.file_data = excel_file

        self.df_mapping = pd.read_excel(self.file_data, sheet_name=self.work_sheet, converters={self.work_codici_disciplina_catalogo: str, self.work_codice_prestazione_siss: str, self.work_codice_metodica: str, self.work_codice_distretto: str}).replace(np.nan, '', regex=True)
        
        
        #print ("print JSON")
        #print(sh)
        #print(checked_dict)
        self.controls_setted = checked_dict #checklist dei test da effettuare
        #self.analizer(df_mapping)'''


    def check_sheet(self, data):
        print("FASE 0: precheck")
        self.check_sheet_existance() #controllo se lo sheet del mapping esiste
        self.check_sheet_fields(data, self.file.df_mapping)
        #if self.validation_results != "": #controlla i nomi dei campi dello sheet
        return self.validation_results
        

    def check_sheet_existance(self):
        print("PRELIMINARE: check the used column name of the excel file")
        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        if self.file.work_sheet in xfile.sheetnames:
            print('sheet exists')
            self.validation_results = ""
        else:
            self.validation_results = self.error_generate_list[1]

    def check_sheet_fields(self, data, df_mapping):
        print("PRELIMINARE: check fields name's of the sheet")
        field_in_error = []
        column_headers = self.file.df_mapping.columns.values.tolist()
        print("The Column Header :", column_headers)

        cont = 0
        for value in data[0]["work_column"].values():
            cont += 1
            #print("value: " + value)
            if value not in column_headers and cont !=1:
                field_in_error.append(value)

        if field_in_error:# != []:
            self.validation_results = self.validation_results + ",\n" + self.error_generate_list[2] + ", ".join(field_in_error)


    def file_validation(self, error_dict):
        print("questi sono gli errori individuati e separati per categoria:\n %s", error_dict)
        #self.output_message = self.output_message + "\n" + json.dumps(error_dict)
        
        print("\nPer osservare i risultati ottenuti, controllare il file prodotto: check_excel_result.txt")
        with open(RESULT_VALIDATION, "w") as f:
            f.write(self.output_message + "\n" + json.dumps(error_dict)) #self.output_message + "\n" + json.dumps(error_dict)
            f.close()
        try:
            #Report_creation = Report_Creation()
            Report_creation = Report_Creation(  self.file.df_mapping,
                                                self.file.file_data,
                                                self.file.work_sheet,
                                                self.file.work_N1,
                                                self.file.work_N2,
                                                self.file.work_descrizione_N1,
                                                self.file.work_descrizione_N2,
                                                self.file.work_codice_prestazione_siss,
                                                self.file.work_descrizione_prestazione_siss,
                                                self.file.work_codice_agenda_siss,
                                                #self.work_casi_1_n,
                                                self.file.work_abilitazione_esposizione_siss,
                                                self.file.work_prenotabile_siss,
                                                self.file.work_codici_disciplina_catalogo,
                                                self.file.work_descrizione_disciplina_catalogo,
                                                self.file.work_codice_QD,
                                                self.file.work_codice_metodica,
                                                self.file.work_codice_distretto,
                                                self.file.work_priorita_U,
                                                self.file.work_priorita_primo_accesso_D,
                                                self.file.work_priorita_primo_accesso_P,
                                                self.file.work_priorita_primo_accesso_B,
                                                self.file.work_accesso_programmabile_ZP,
                                                self.file.work_combinata,
                                                self.file.work_codice_agenda_interno,
                                                self.file.work_codice_prestazione_interno,
                                                self.file.work_inviante,
                                                self.file.work_accesso_farmacia,
                                                self.file.work_accesso_CCR,
                                                self.file.work_accesso_cittadino,
                                                self.file.work_accesso_MMG,
                                                self.file.work_accesso_amministrativo,
                                                self.file.work_accesso_PAI,
                                                self.file.work_gg_preparazione,
                                                self.file.work_gg_refertazione,
                                                self.file.work_nota_operatore,
                                                self.file.work_nota_agenda,
                                                self.file.work_nota_revoca,
                                                self.file.work_disciplina,
                                                self.file.work_sesso,
                                                self.file.work_eta_min,
                                                self.file.work_eta_max,
                                                self.file.work_alert_column,
                                                self.file.work_delimiter,
                                                error_dict)
            Report_creation.get_report()

        except Exception as e:
            logger.error('Failed to upload to ftp: '+ str(e))
            print("Errore in Report Creation")
        '''try:
            xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
            #sheet_mapping = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel mapping
            try:
                sheet_report = xfile.get_sheet_by_name('Report Validazione') #recupero sheet excel report validazione 
            except:
                #creo Report Validazione se non esiste
                sheet_report = xfile.create_sheet('Report Validazione')
            sheet_report["A1"] = "Report Validazione"

            xfile.save(self.file_data) 
        except:
            print("non esiste file o sheet")'''
        
        

    def findCell(self, sh, searchedValue, start_col):
        result_coord = []
        result_value = []

        for row in range(sh.nrows):
            for col in range(start_col, start_col+1):
                myCell = sh.cell(row, col)
                myValue = sh.cell(row, self.file.work_index_codice_prestazione_SISS)
                abilita = sh.cell(row, self.file.work_index_abilitazione_esposizione_SISS)
                if myCell.value == searchedValue and abilita.value == "S":
                    result_coord.append(str(row) + "#" + str(col))
                    result_value.append(myValue.value)
                    #return row, col#xl_rowcol_to_cell(row, col)

        if result_coord == []:
            return -1
        return result_coord#, result_value

    def findCell_agenda(self, sh, searchedValue, start_col):
        result_coord = []
        result_value = []

        for row in range(sh.nrows):
            for col in range(start_col, start_col+1):
                myCell = sh.cell(row, col)
                myValue = sh.cell(row, self.file.work_index_codice_SISS_agenda) #Codice SISS agenda 15
                #abilita = sh.cell(row, self.work_index_abilitazione_esposizione_SISS-1) #abilitazione esposizione SISS 28
                if myCell.value == searchedValue: # and abilita.value == "S":
                    result_coord.append(str(row) + "|" + str(col))
                    result_value.append(myValue.value)
                    #return row, col#xl_rowcol_to_cell(row, col)

        if result_coord == []:
            return -1
        return result_coord#, result_value
                                    
    def findCell_agenda_II(self, df, searchedValue, name_column):#sheet_mapping, searchedAgenda, self.work_index_codice_SISS_agenda
        result_coord = []

        for rowIndex, row in df.iterrows(): #iterate over rows
            myCell = row[name_column]
            abilita = row[self.file.work_abilitazione_esposizione_siss]
            if myCell == searchedValue and abilita == "S":
                result_coord.append(str(rowIndex))

        if result_coord == []:
            return -1
        return result_coord

    '''Metodo che aggiunge elemento in una lista esistente o crea la lista nel caso
    non fosse presente'''
    def update_list_in_dict(self, dictio, index, element):
        if index in dictio.keys():
            dictio[index].append(element)
        else:
            dictio[index] = [element]
        return dictio

    '''Metodo per controllare il tipo di una colonna excel'''
    def column_validator(self):
        return ""

    def list_duplicates(self, seq):
        seen = set()
        seen_add = seen.add
        # adds all elements it doesn't know yet to seen and all other to seen_twice
        seen_twice = set( x for x in seq if x in seen or seen_add(x) )
        # turn the set into a list (as requested)
        return list( seen_twice )