import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import yaml
import logging
import re
import openpyxl
from flaskr.check_action import Check_action

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

class Check_canali(Check_action):

    output_message = ""
    error_list = {}

    def __init__(self, file):
        #pass
        super().__init__(file)
        self.file = file

   
    def ck_canali_vuoti(self, error_dict):
        print("start checking canali di prenotazione configurati")
        error_dict.update({'error_canali_vuoti': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel
        
        type_error = {
            1: "farmacia",
            2: "ccr",
            3: "cittadino",
            4: "mmg",
            5: "amministrativo",
            6: "pai"
        }
        list_error = []

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S" and row[self.file.work_prenotabile_siss] == "S": #se prestazione esposta e prenotabile
                if row[self.file.work_accesso_farmacia] == "":
                    list_error.append(type_error[1])
                if row[self.file.work_accesso_CCR] == "":
                    list_error.append(type_error[2])
                if row[self.file.work_accesso_cittadino] == "":
                    list_error.append(type_error[3])
                if row[self.file.work_accesso_MMG] == "":
                    list_error.append(type_error[4])
                if row[self.file.work_accesso_amministrativo] == "":
                    list_error.append(type_error[5])
                if row[self.file.work_accesso_PAI] == "":
                    list_error.append(type_error[6])

                if list_error != []:
                    error_dict["error_canali_vuoti"].append(str(int(index)+2))
                    ind = str(int(index)+2)
                    out_message = ""
                    out_message = "__> I seguenti canali di accesso non sono valorizzati: '{}' ".format(", ".join(list_error))
                    if sheet[self.file.work_alert_column+ind].value is not None:
                        sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message
                    else:
                        sheet[self.file.work_alert_column+ind] = out_message

            list_error = []

        xfile.save(self.file.file_data)  
        return error_dict

    def ck_canali_PAI(self, error_dict):
        print("start checking canale PAI")
        error_dict.update({'error_canale_PAI': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S" and row[self.file.work_prenotabile_siss] == "S": #se prestazione esposta e prenotabile
                if row[self.file.work_accesso_PAI] == "S":
                    if row[self.file.work_accesso_farmacia] == "S" or row[self.file.work_accesso_CCR] == "S" or row[self.file.work_accesso_cittadino] == "S" or row[self.file.work_accesso_MMG] == "S" or row[self.file.work_accesso_amministrativo] == "S":
                        error_dict["error_canale_PAI"].append(str(int(index)+2))
            
        out_message = ""
        for ind in error_dict['error_canale_PAI']:
            out_message = "__> Rilevati canali di accesso abilitati contemporaneamente al canale CREG PAI"
            out_message = out_message + "\n _> lasciare abilitato solo il canale CREG PAI o disabilitarlo."
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)  
        return error_dict

    def ck_canali_abilitati(self, error_dict):
        print("start checking canali abilitati")
        error_dict.update({'error_canali_abilitati': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S" and row[self.file.work_prenotabile_siss] == "S": #se prestazione esposta e prenotabile
                if row[self.file.work_accesso_PAI] == "S":
                    if row[self.file.work_accesso_farmacia] == "N" and row[self.file.work_accesso_CCR] == "N" and row[self.file.work_accesso_cittadino] == "N" and row[self.file.work_accesso_MMG] == "N" and row[self.file.work_accesso_amministrativo] == "N" and row[self.file.work_accesso_PAI] == "N":
                        error_dict["error_canali_abilitati"].append(str(int(index)+2))
            
        out_message = ""
        for ind in error_dict['error_canali_abilitati']:
            out_message = "__> Rilevati canali di accesso tutti disabilitati per prestazione esposta e prenotabile"
            out_message = out_message + "\n _> abilitare almeno uno dei canali di prenotazione proposti."
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)
        return error_dict