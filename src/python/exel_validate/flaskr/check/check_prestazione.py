import pandas as pd
import numpy as np
from xlsxwriter.utility import xl_rowcol_to_cell
import logging
import openpyxl
from flaskr.check_action import Check_action

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

class Check_prestazione(Check_action):

    output_message = ""
    error_list = {}

    def __init__(self, file):
        #pass
        super().__init__(file)
        self.file = file
    
    def ck_casi_1n(self, error_dict):
        print("start checking if casi 1:n is correct")
        error_dict.update({'error_casi_1n': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        casi_1n_dict_error = {}
        metodica_distretti_dict = {} #dict delle metodiche e distretti delle prestazioni messe in lista
        distretti_dict = {}
        index_dict = {}
        
        for index, row in self.file.df_mapping.iterrows():
            a_p = str(row[self.file.work_codice_agenda_siss]) + "|" + str(row[self.file.work_codice_prestazione_siss])
            m_d = str(row[self.file.work_codice_metodica]) + "|" + str(row[self.file.work_codice_distretto])
            index_dict = self.update_list_in_dict(index_dict, a_p, str(int(index)+2))
            distretti_dict = self.update_list_in_dict(distretti_dict, str(int(index)+2), row[self.file.work_operatore_logico_distretto])
            if row[self.file.work_codice_prestazione_siss] != "" and row[self.file.work_codice_agenda_siss] != "":
                if row[self.file.work_abilitazione_esposizione_siss] == "S":
                    if a_p not in metodica_distretti_dict.keys():
                        #primo elemento inserito nel dict
                        metodica_distretti_dict = self.update_list_in_dict(metodica_distretti_dict, a_p, m_d)
                    elif a_p in metodica_distretti_dict.keys():
                        #secondo e successivi elementi inseriti nel dict
                        if m_d not in metodica_distretti_dict[a_p]:
                            #elemento non costituisce caso 1:N al momento
                            metodica_distretti_dict = self.update_list_in_dict(metodica_distretti_dict, a_p, m_d)
                        elif m_d in metodica_distretti_dict[a_p]:
                            #elemento costituisce caso 1:N. 2 o più elementi con stesso m_d
                            metodica_distretti_dict = self.update_list_in_dict(metodica_distretti_dict, a_p, m_d)
                            #error_dict = self.update_list_in_dict(error_dict, str(int(index)+2), a_p)
                        else:
                            #teoricamente non dovrebbe mai entrare in questa condizione
                            metodica_distretti_dict = self.update_list_in_dict(metodica_distretti_dict, a_p, m_d)

        #md_dict_total = metodica_distretti_dict
        print("valutazione risultati casi 1:N")
        for key, value in metodica_distretti_dict.items(): #key:a_p, value:m_d
            print(key + ":  " + ", ".join(value))
            indexAP = index_dict[key] #indici dove si trovano tutte le occorrenze di una coppia A/P
            lengthMD = len(value) #occorrenze coppie prestazioni/agenda

            flag_error1 = False
            flag_error2 = False
            flag_error3 = False

            '''error_distretti_duplicati = get_distretti_duplicates_in_agenda(self, value)
            if error_distretti_duplicati == -1:
                flag_error1 = True #da modificare'''

            set_duplicates = {} #dict con key gli m_d e value sono le occorrenze
            if lengthMD > 1: #occorrenza multipla, rilevo possibile caso 1:N se lenght > 1
                print("occorrenza multipla " + str(lengthMD))
                
                error_1_list = [] 
                error_2_list = [] 
                cont = 0
                dict_items_cont = {} #inserisco chiave:m_d e value: a che indice di indexMD è
                distretti_list = []

                for v in value: #per ogni m_d in value
                    print("v: " + v)
                    #definisco un dict per conteggiare le occorrenze degli m_d
                    if v in set_duplicates.keys(): 
                        set_duplicates[v] = set_duplicates[v] + 1
                        dict_items_cont[v].append(cont)
                        #print("1:" + str(set_duplicates[v]))
                    else:
                        set_duplicates[v] = 1
                        dict_items_cont[v] = [cont]
                        #print("2:" + str(set_duplicates[v]))
                    
                    splitMD_list = v.split("|") #splitto m_d in una lista di due elementi
                    print("splitMD_list: " + splitMD_list[1])
                    split_district = splitMD_list[1]
                    #verifico se nei casi 1:N, i distretto non sono vuoti
                    #if split_district != '': #and self.list_duplicates(splitMD_list[0].split(self.work_delimiter)) != []: 
                        #contM = 0
                    if split_district == "": #errore se il distretto è vuoto
                        print("d: " + split_district)
                        if split_district == "": 
                            flag_error1 = True
                            error_1_list.append(indexAP[cont])
                    else: #errore se i distretti usati sono ripetuti su più coppie
                        for distretto in split_district.split(self.file.work_delimiter):
                            if distretto in distretti_list and flag_error2 == False: 
                                flag_error2 = True
                                error_2_list.append(indexAP[cont])
                            distretti_list.append(distretto)
                    print("distretti_list: ", distretti_list)        
                    cont = cont + 1
                
                #verifico le occorrenze degli m_d
                #se le occorrenze di un m_d sono > 1, allora caso 1:N non risolto
                for MD, occorrenze in set_duplicates.items(): 
                    if occorrenze > 1: #se le occorrenze sono > 1 allora caso 1:N
                        flag_error2 = True
                        for md in dict_items_cont[MD]:
                            if indexAP[md] not in error_2_list: 
                                error_2_list.append(indexAP[md])
                
                #se flag_error1 è True, allora trovato errore distretto vuoto
                if flag_error1 == True: 
                    for ind in error_1_list:
                        print("flag error 1: " + ind)
                        casi_1n_dict_error = self.update_list_in_dict(casi_1n_dict_error, ind, key + ": caso 1:N con distretto vuoto")
                        if ind not in error_dict["error_casi_1n"]:
                            error_dict["error_casi_1n"].append(ind) #ind ha già sommato + 2
                #se flag_error2 è True, allora trovato errore chiave m_d multipla
                if flag_error2 is True: 
                    for ind in error_2_list:
                        print("flag error 2: " + ind)
                        fl_value_null = False
                        for v in value:
                            if v != "|": #se true allora ci sono metodiche o distretti
                                fl_value_null = True
                        if fl_value_null is True:
                            casi_1n_dict_error = self.update_list_in_dict(casi_1n_dict_error, ind, key + " sono presenti delle metodiche e/o distretti non univoci che per specializzare la coppia prestazione/agenda: " + ", ".join(value))
                        else:
                            casi_1n_dict_error = self.update_list_in_dict(casi_1n_dict_error, ind, key + " le metodiche e distretti non sono stati valorizzati per risolvere caso 1:N")
                        if ind not in error_dict["error_casi_1n"]:
                            error_dict["error_casi_1n"].append(ind) #ind ha già sommato + 2
                         
                #verifico se gli operatori logici sono conformi e univoci per ogni coppia A/P nei casi 1:N
                opd = "U"
                for ind in indexAP:
                    distretto_op = distretti_dict[ind] 
                    if opd != distretto_op[0] and self.file.df_mapping.at[int(ind)-2,self.file.work_codice_distretto] != "":
                        print("dis:" + self.file.df_mapping.at[int(ind)-2,self.file.work_codice_distretto])
                        flag_error3 = True
                        print("flag error 4: " + ind)
                        if ind not in error_dict["error_casi_1n"]:
                            error_dict["error_casi_1n"].append(ind) #ind ha già sommato + 2
                        casi_1n_dict_error = self.update_list_in_dict(casi_1n_dict_error, ind, key + ": Operatore Logico Distretti non conforme all'interno della coppia agenda/prestazione")
                    #opd = distretto_op[0]
    
                #taggo i casi 1:N risolti, cioè che non presentano errori
                for i in indexAP:
                    out_message = ""
                    if flag_error1 is False and flag_error2 is False and flag_error3 is False:
                        out_message = "__> Caso 1:N:\n"
                        out_message = out_message + "  _> risolto "
                        if sheet[self.file.work_alert_column+i].value is not None:
                            sheet[self.file.work_alert_column+i] = str(sheet[self.file.work_alert_column+i].value) + "; \n" + out_message #modificare colonna alert
                        else:
                            sheet[self.file.work_alert_column+i] = out_message
                    elif i not in error_dict["error_casi_1n"]:
                        out_message = "__> Caso 1:N:\n"
                        out_message = out_message + "  _> distretti o metodiche coerenti"
                        if sheet[self.file.work_alert_column+i].value is not None:
                            sheet[self.file.work_alert_column+i] = str(sheet[self.file.work_alert_column+i].value) + "; \n" + out_message #modificare colonna alert
                        else:
                            sheet[self.file.work_alert_column+i] = out_message

                    

        #creazione del messaggio di alert riportato nel file excel
        print("start definizione output casi 1:n")
        out1 = ""
        out_message = ""
        for ind in error_dict['error_casi_1n']:
            out_message = "__> Caso 1:N:\n"
            out_message = out_message + "  _> Per la coppia agenda/prestazione: '{}'".format(", ".join(casi_1n_dict_error[ind]))
            #if "S" in df_mapping.at[int(ind)-2, self.file.work_combinata] or "1" in df_mapping.at[int(ind)-2, self.file.work_combinata] or "2" in df_mapping.at[int(ind)-2, self.file.work_combinata]: #verifico se c'è combinata
            #    out_message = out_message + ";\n  _> rilevata possibile risoluzione tramite combinata"
            out1 = out1 + "at index: " + ind + ", on agenda_prestazione: " + ", ".join(casi_1n_dict_error[ind]) + ", \n"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        self.output_message = self.output_message + "\nerror_casi_1n: \n" + "at index: \n" + out1
            
        xfile.save(self.file.file_data)
        return error_dict

    #per ogni agenda, segnala le prestazioni senza codice prestazione SISS associato
    def ck_prestazione(self, error_dict):
            
        print("start checking if prestazione has the associated codice prestazione SISS")
        error_dict.update({'error_prestazione_SISS_assente': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        
        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_codice_prestazione_siss] == "" and row[self.file.work_codice_agenda_siss] != "":
                if row[self.file.work_abilitazione_esposizione_siss] == "S":
                    error_dict["error_prestazione_SISS_assente"].append(str(int(index)+2))

        out_message = ""
        for ind in error_dict['error_prestazione_SISS_assente']:
            out_message = "__> Codice Prestazione SISS assente: per la seguente prestazione abilitata all'esposizione SISS non è presente il codice prestazione SISS."
            out_message = out_message + "\n _> Nel caso si volesse esporre la prestazione, inserire il codice prestazione SISS, altrimenti disabilitate la prestazione all'esposizione" 
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)
        return error_dict

    #controlla se la nota operatore è stata configurata per le prestazioni esposte ma non prenotabili
    def ck_prestazione_nonprenotabile(self, error_dict):
            
        print("start checking if prestazione non prenotabile has the nota operativa configured")
        error_dict.update({
            'error_notaop_assente': [],
            'error_prenotabileflag_assente': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        
        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                if row[self.file.work_prenotabile_siss] == "":
                    error_dict["error_prenotabileflag_assente"].append(str(int(index)+2))
                elif row[self.file.work_prenotabile_siss] == "N" and row[self.file.work_nota_operatore] == "":
                    error_dict["error_notaop_assente"].append(str(int(index)+2))
        
        out_message = ""
        for ind in error_dict['error_prenotabileflag_assente']:
            out_message = "__> flag prenotabile assente: per la seguente prestazione abilitata all'esposizione SISS, inserire il flag prenotabilita a 'S' o 'N'"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message 
        for ind in error_dict['error_notaop_assente']:
            out_message = "__> Nota operatore assente: per la seguente prestazione abilitata all'esposizione SISS ma non prenotabile, è necessario inserire la nota operatore."
            out_message = out_message + "\n _> Indicare nella nota operatore se il cittadino dovrà contattare direttamente la struttura per prenotare o indicare altre motivazioni" 
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)
        return error_dict