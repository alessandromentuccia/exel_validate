import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import yaml
import logging
import re
import openpyxl
from flaskr.check_action import Check_action

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

class Check_metodiche(Check_action):

    #file_name = ""
    output_message = ""
    error_list = {}

    work_index_codice_QD = 0
    work_index_codice_SISS_agenda = 0
    work_index_abilitazione_esposizione_SISS = 0
    work_index_codice_prestazione_SISS = 0
    work_index_operatore_logico_distretto = 0
    work_index_codici_disciplina_catalogo = 0

    def __init__(self, file):
        super().__init__(file)
        self.file = file

    
    def ck_metodica_inprestazione(self, error_dict):
        print("start checking if metodica are correct")
        error_dict.update({'error_metodica_inprestazione': [],
                           'error_metodica_generale_assente': [],
                           'error_metodica_generale_presente': []})
        #self.output_message = self.output_message + "\nerror_metodica_inprestazione: "
        
        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel
        
        cod_metodica_generale = "M00000"

        metodica_dict_error = {}
        prestazioni_dict = {}
        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                Metodica_string = row[self.file.work_codice_metodica]
                Metodica_list = Metodica_string.split(self.file.work_delimiter)


                siss = "" 
                siss_flag = False
                if Metodica_list is not None:

                    cod_pre_siss = str(row[self.file.work_codice_prestazione_siss]).strip()
                    if cod_metodica_generale == Metodica_string.strip():
                        error_dict['error_metodica_generale_presente'].append(str(int(index)+2))
                    else:
                        if cod_metodica_generale not in Metodica_string and Metodica_string.strip() != "":
                            error_dict['error_metodica_generale_assente'].append(str(int(index)+2))
                            #print("codice metodica: '" + str(Metodica_string).strip() + "'")
                            
                        for metodica in Metodica_list:
                            metodica = metodica.strip()
                            if metodica != "":
                                short_sheet = self.file.sheet_Metodiche.loc[self.file.sheet_Metodiche["Codice Metodica"] == metodica]                      
                                
                                print("prestazione della metodica in mapping:" + cod_pre_siss + " + " + siss)
                                
                                if cod_pre_siss not in short_sheet["Codice SISS"].values:
                                    siss_flag = True
                                    metodica_dict_error = self.update_list_in_dict(metodica_dict_error, str(int(index)+2), metodica)
                                    prestazioni_dict[str(int(index)+2)] = cod_pre_siss
                                    print("error metodica on index:" + str(int(index)+2))
                                else:
                                    if cod_pre_siss != siss and siss != "":
                                        print("error metodica on index:" + str(int(index)+2))
                                        siss_flag = True
                                        metodica_dict_error = self.update_list_in_dict(metodica_dict_error, str(int(index)+2), metodica)
                                        prestazioni_dict[str(int(index)+2)] = cod_pre_siss
                if siss_flag == True: #se durante il mapping con la sua disciplina, questa non viene rilevata, allora è errore
                    error_dict['error_metodica_inprestazione'].append(str(int(index)+2))
                siss = str(row[self.file.work_codice_prestazione_siss]).strip()

            
        out1 = ""
        out_message = ""
        for ind in error_dict['error_metodica_inprestazione']:
            out1 = out1 + "at index: " + ind + ", on metodica: " + ", ".join(metodica_dict_error[ind]) + ", \n"
            out_message = "__> Metodiche: '{}' non previste per la prestazione: '{}'".format(", ".join(metodica_dict_error[ind]), prestazioni_dict[ind])
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        for ind in error_dict['error_metodica_generale_assente']:
            out1 = out1 + "at index: " + ind + ", metodica generale assente \n"
            out_message = "__> Alert: si consiglia di verificare la METODICA GENERALE non configurata per la prestazione"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        for ind in error_dict['error_metodica_generale_presente']:
            out1 = out1 + "at index: " + ind + ", metodica generale presente \n"
            out_message = "__> Alert: METODICA GENERALE configurata da sola - Si consiglia di lasciare vuoto il campo poichè è sottintesa"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        

        self.output_message = self.output_message + "\nerror_metodica_inprestazione: \n" + out1

        

        xfile.save(self.file.file_data)
        return error_dict

    def ck_metodica_sintassi(self, error_dict):
        print("start checking if there is ',' separator between each metodiche defined")
        error_dict.update({
            'error_metodica_caratteri_non_consentiti': [],
            'error_metodica_spazio_bordi': [],
            'error_metodica_spazio_internamente': []
        })
        string_check = re.compile('1234567890,M')

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                print("Metodica: " + row[self.file.work_codice_metodica])
                #flag_error = False
                if row[self.file.work_codice_metodica] is not None:
                    
                    #row_replace = row[self.file.work_codice_metodica].replace(" ", "")
                    if " " in row[self.file.work_codice_metodica]:
                        if row[self.file.work_codice_metodica] != row[self.file.work_codice_metodica].strip():
                            print("string contain space in the border")
                            error_dict['error_metodica_spazio_bordi'].append(str(int(index)+2))
                        r = row[self.file.work_codice_metodica].strip()
                        if " " in r:
                            print("string contain space inside the string")
                            error_dict['error_metodica_spazio_internamente'].append(str(int(index)+2))
                        row_replace = row[self.file.work_codice_metodica].replace(" ", "")
                        if(string_check.search(row_replace) is not None): #!=
                            print("String contains other Characters.")
                            error_dict['error_metodica_caratteri_non_consentiti'].append(str(int(index)+2))

                    '''if " " in row[self.file.work_codice_metodica]:
                        if " " in row_replace.strip():
                            print("string contain space inside the string")
                            error_dict['error_metodica_spazio_internamente'].append(str(int(index)+2))
                        else:
                            print("string contain space in the border")
                            error_dict['error_metodica_spazio_bordi'].append(str(int(index)+2))
                    elif(string_check.search(row_replace) != None):
                        print("String contains other Characters.")
                        error_dict['error_metodica_caratteri_non_consentiti'].append(str(int(index)+2))'''


        #da eliminare start
        out1 = ", \n".join(error_dict['error_metodica_caratteri_non_consentiti'])
        self.output_message = self.output_message + "\nerror_metodica_caratteri_non_consentiti: \n" + "at index: \n" + out1
        out2 = ", \n".join(error_dict['error_metodica_spazio_bordi'])
        self.output_message = self.output_message + "\nerror_metodica_spazio_bordi: \n" + "at index: \n" + out2
        out3 = ", \n".join(error_dict['error_metodica_spazio_internamente'])
        self.output_message = self.output_message + "\nerror_metodica_spazio_internamente: \n" + "at index: \n" + out3
        #stop

        out_message = ""
        for ind in error_dict['error_metodica_caratteri_non_consentiti']:
            out_message = "__> Metodiche presentano errori di sintassi: rilevati caratteri non consentiti"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        for ind in error_dict['error_metodica_spazio_bordi']:
            out_message = "__> Metodiche presentano errori di sintassi: rilevati degli spazi alle estremità del contenuto della cella"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        for ind in error_dict['error_metodica_spazio_internamente']:
            out_message = "__> Metodiche presentano errori di sintassi: rilevati degli spazi all'interno del contenuto della cella"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data) 
        return error_dict #da eliminare
    
    def ck_metodica_descrizione(self, error_dict):
        print("start checking if metodica have the correct description")
        error_dict.update({
            'error_metodica_descrizione': [],
            'error_metodica_separatore': []
            })

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        metodica_dict_error = {}
        
        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                Metodica_string = row[self.file.work_codice_metodica].split(self.file.work_delimiter)
                Description_list = row[self.file.work_descrizione_metodica].split(self.file.work_delimiter)
                flag_error = False
                
                if Metodica_string is not None:
                    for metodica in Metodica_string:
                        metodica = metodica.strip()
                        if metodica != "":
                            metodica_catalogo = self.file.sheet_Metodiche.loc[self.file.sheet_Metodiche["Codice Metodica"] == metodica]                    
                            
                            try:
                                if metodica_catalogo["Metodica Rilevata"].values[0] not in Description_list:
                                    print("la descrizione metodica non è presente all'indice " + str(int(index)+2))
                                    flag_error = True
                                    metodica_dict_error = self.update_list_in_dict(metodica_dict_error, str(int(index)+2), metodica)
                            except:
                                if metodica_catalogo.size == 0:
                                    #print("print metodica_catalogo2:" + metodica_catalogo)
                                    print("exception avvennuta: controllare manualmente qual'è il problema")
                                    logging.error("controllare manualmente la metodica: " + metodica + " all'indice: " + str(int(index)+2))
                                    flag_error = True
                                    metodica_dict_error = self.update_list_in_dict(metodica_dict_error, str(int(index)+2), metodica)
                if flag_error == True:
                    error_dict['error_metodica_descrizione'].append(str(int(index)+2))

        out1 = ""
        out_message = ""
        for ind in error_dict['error_metodica_descrizione']:
            out1 = out1 + "at index: " + ind + ", on metodica: " + ", ".join(metodica_dict_error[ind]) + ", \n"
            out_message = "__> Metodiche: '{}' presentano errori nella descrizione".format(", ".join(metodica_dict_error[ind]))
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        self.output_message = self.output_message + "\nerror_metodica_descrizione: \n" + out1
        for ind in error_dict['error_metodica_separatore']:
            out_message = "__> Le descrizioni metodiche presentano errori col separatore ','"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        out2 = ", \n".join(error_dict['error_metodica_separatore'])
        self.output_message = self.output_message + "\nerror_metodica_separatore: \n" + out2

        xfile.save(self.file.file_data)
        return error_dict
