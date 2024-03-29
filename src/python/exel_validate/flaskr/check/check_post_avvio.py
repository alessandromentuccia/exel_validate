import os
from dotenv import load_dotenv
import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import yaml
import logging
import re
import openpyxl 

load_dotenv()
ALERT_COLUMN_RV=os.getenv("ALERT_COLUMN_RV")
MAP_VALUE_COLUMN_RV=os.getenv("MAP_VALUE_COLUMN_RV")

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

'''La seguente classe permette di gestire i controlli che vengono effettuati tra rivisto e mapping e sui campi che vengono selezionati nel form'''
class Check_post_avvio():
    '''Il metodo rileva i controlli che sono stati selezionati nella form e li manda in esecuzione, sia su rivisto che su mapping'''
    def ck_post_avvio(self, df_mapping, df_rivisto, error_dict):
        print("start checking post avvio")
        error_dict.update({'error_post_avvio': []})

        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet) #recupero sheet excel
 
        if self.configurazione_rivisto["Quesiti"] != "": #check list
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Quesiti", self.work_codice_QD, "list")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Quesiti", self.work_codice_QD, "list")
            #error_dict = Check_post_avvio.ck_QD(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["OperatoreQD"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "OperatoreQD", self.work_operatore_logico_QD, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "OperatoreQD", self.work_operatore_logico_QD, "string")
            #error_dict = Check_post_avvio.ck_Operatore_QD(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Distretti"] != "": #check list
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Distretti", self.work_codice_distretto, "list")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Distretti", self.work_codice_distretto, "list")
            #error_dict = Check_post_avvio.ck_Distretti(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["OperatoreDistretto"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "OperatoreDistretto", self.work_operatore_logico_distretto, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "OperatoreDistretto", self.work_operatore_logico_distretto, "string")
            #error_dict = Check_post_avvio.ck_Operatore_Distretti(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Metodiche"] != "": #check list
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Metodiche", self.work_codice_metodica, "list")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Metodiche", self.work_codice_metodica, "list")
            #error_dict = Check_post_avvio.ck_Metodiche(self, df_mapping, df_rivisto, error_dict)    
        if self.configurazione_rivisto["Inviante"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Inviante", self.work_inviante, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Inviante", self.work_inviante, "string")
            #error_dict = Check_post_avvio.ck_Invianti(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Risorsa"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Risorsa", self.work_risorsa, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Risorsa", self.work_risorsa, "string")
            #error_dict = Check_post_avvio.ck_Risorsa(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Farmacia"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Farmacia", self.work_accesso_farmacia, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Farmacia", self.work_accesso_farmacia, "string")
            #error_dict = Check_post_avvio.ck_Farmacia(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["CCR"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "CCR", self.work_accesso_CCR, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "CCR", self.work_accesso_CCR, "string")
            #error_dict = Check_post_avvio.ck_CCR(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Cittadino"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Cittadino", self.work_accesso_cittadino, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Cittadino", self.work_accesso_cittadino, "string")
            #error_dict = Check_post_avvio.ck_Cittadino(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["MMG"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "MMG", self.work_accesso_MMG, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "MMG", self.work_accesso_MMG, "string")
            #error_dict = Check_post_avvio.ck_MMG(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["Amministrativo"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "Amministrativo", self.work_accesso_amministrativo, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "Amministrativo", self.work_accesso_amministrativo, "string")
            #error_dict = Check_post_avvio.ck_Amministrativo(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["PAI"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "PAI", self.work_accesso_PAI, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "PAI", self.work_accesso_PAI, "string")
            #error_dict = Check_post_avvio.ck_PAI(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["NoteOperatore"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "NoteOperatore", self.work_nota_operatore, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "NoteOperatore", self.work_nota_operatore, "string")
            #error_dict = Check_post_avvio.ck_NoteOperatore(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["NotePreparazione"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "NotePreparazione", self.work_nota_preparazione, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "NotePreparazione", self.work_nota_preparazione, "string")
            #error_dict = Check_post_avvio.ck_NotePreparazione(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["NoteAmministrative"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "NoteAmministrative", self.work_nota_agenda, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "NoteAmministrative", self.work_nota_agenda, "string")
            #error_dict = Check_post_avvio.ck_NoteAmministrative(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["NoteRevoca"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "NoteRevoca", self.work_nota_revoca, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "NoteRevoca", self.work_nota_revoca, "string")
            #error_dict = Check_post_avvio.ck_NoteRevoca(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["PrioritaUrgenza"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "PrioritaUrgenza", self.work_priorita_U, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "PrioritaUrgenza", self.work_priorita_U, "string")
            #error_dict = Check_post_avvio.ck_PrioritaUrgenza(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["PrioritaOB"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "PrioritaOB", self.work_priorita_primo_accesso_B, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "PrioritaOB", self.work_priorita_primo_accesso_B, "string")
            #error_dict = Check_post_avvio.ck_PrioritaOB(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["PrioritaOD"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "PrioritaOD", self.work_priorita_primo_accesso_D, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "PrioritaOD", self.work_priorita_primo_accesso_D, "string")
            #error_dict = Check_post_avvio.ck_PrioritaOD(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["PrioritaOP"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "PrioritaOP", self.work_priorita_primo_accesso_P, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "PrioritaOP", self.work_priorita_primo_accesso_P, "string")
            #error_dict = Check_post_avvio.ck_PrioritaOP(self, df_mapping, df_rivisto, error_dict)
        if self.configurazione_rivisto["AccessoProgrammabile"] != "":
            error_dict = Check_post_avvio.ck_MAP(self, df_mapping, df_rivisto, error_dict, "AccessoProgrammabile", self.work_accesso_programmabile_ZP, "string")
            error_dict = Check_post_avvio.ck_RIV(self, df_mapping, df_rivisto, error_dict, "AccessoProgrammabile", self.work_accesso_programmabile_ZP, "string")
            #error_dict = Check_post_avvio.ck_AccessoProgrammabile(self, df_mapping, df_rivisto, error_dict)



        #creazione del messaggio di alert riportato nel file excel
        print("start definizione output controlli post avvio")
        return error_dict

    '''Il metodo controlla se gli attributi del rivisto sono configurati allo stesso modo anche nel mapping,
    il risultato sarà osservabile nel rivisto'''
    def ck_MAP(self, df_mapping, df_rivisto, error_dict, element, work_codice, flag_type):
        print("start checking campo Rivisto in MAPPING")
        error_dict.update({
            'error_ck_'+element: [] })

        alert_column = ALERT_COLUMN_RV
        map_value_column = MAP_VALUE_COLUMN_RV

        xfile = openpyxl.load_workbook(self.file_rivisto) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.configurazione_rivisto["Sheet"])

        file_rivisto = self.file_rivisto 
        configurazione_rivisto = self.configurazione_rivisto

        agenda = configurazione_rivisto["Agenda"] #intesta
        prestazioneSISS = configurazione_rivisto["PrestazioneSISS"]
        prestazioneInterna = configurazione_rivisto["PrestazioneInterna"]

        for index, row in df_rivisto.iterrows():
            rivisto_key = str(row[agenda]).strip()+"|"+str(row[prestazioneSISS]).strip()+"|"+str(row[prestazioneInterna]).strip()
            #print("iterate row: " + rivisto_key)
            #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
            searchedValue = row[configurazione_rivisto[element]] 
            if element == "Inviante":
                if searchedValue == "":
                    searchedValue = "0"

            if flag_type == "list":
                result_ = self.findCell_dataframe_MAP(df_mapping, searchedValue, rivisto_key, work_codice)
            elif flag_type == "string":
                result_ = self.findCell_dataframe_MAP_string(df_mapping, searchedValue, rivisto_key, work_codice) 
            
            if result_ == -2:
                error_dict["error_ck_"+element].append(str(int(index)+2))
                print("coppia non trovata nel mapping: " +element)
                out_message = "__> {}".format("Coppia prestazione/agenda non trovata in mapping")
                if sheet[alert_column+str(int(index)+2)].value is not None:
                    sheet[alert_column+str(int(index)+2)] = str(sheet[alert_column+str(int(index)+2)].value) + "; \n" + out_message 
                else:
                    sheet[alert_column+str(int(index)+2)] = out_message
            elif isinstance(result_, str):
                error_dict["error_ck_"+element].append(str(int(index)+2))
                print("trovato errore su "+element+ " : " + result_)
                out_message = "__> {}".format("Corrispondenza "+element+" non trovata in mapping")
                if sheet[alert_column+str(int(index)+2)].value is not None:
                    sheet[alert_column+str(int(index)+2)] = str(sheet[alert_column+str(int(index)+2)].value) + "; \n" + out_message 
                else:
                    sheet[alert_column+str(int(index)+2)] = out_message
                if sheet[map_value_column+str(int(index)+2)].value is not None:
                    sheet[map_value_column+str(int(index)+2)] = str(sheet[map_value_column+str(int(index)+2)].value) + "; \n" + result_ #modificare result_
                else:
                    sheet[map_value_column+str(int(index)+2)] = result_ #modificare result_   
            elif result_ == -1:
                print("trovata corrispondenza "+element)
                '''out_message = "__> {}".format(element + " corrisponde in mapping")
                if sheet[alert_column+str(int(index)+2)].value is not None:
                    sheet[alert_column+str(int(index)+2)] = str(sheet[alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                else:
                    sheet[alert_column+str(int(index)+2)] = out_message'''
                
        xfile.save(self.file_rivisto)     
        return error_dict
    
    '''Metodo che controlla se gli attributi del mapping sono anche sul rivisto,
        il risultato sarà osservabile nel mapping'''
    def ck_RIV(self, df_mapping, df_rivisto, error_dict, element, work_codice, flag_type):
        print("start checking campo Mapping in RIVISTO")
        error_dict.update({
            'error_ck_reverse_'+element: [] })
        
        xfile = openpyxl.load_workbook(self.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.work_sheet)
        
        file_rivisto = self.file_rivisto  
        configurazione_rivisto = self.configurazione_rivisto

        agenda = self.work_codice_agenda_siss #intesta
        prestazioneSISS = self.work_codice_prestazione_siss
        prestazioneInterna = self.work_codice_prestazione_interno

        for index, row in df_mapping.iterrows():
            if row[self.work_abilitazione_esposizione_siss] == "S":
                mapping_key = str(row[agenda]).strip()+"|"+str(row[prestazioneSISS]).strip()+"|"+str(row[prestazioneInterna]).strip()
                #print("iterate row: " + rivisto_key)
                #print("iterate row: " + row[self.configurazione_rivisto["Quesiti"]])
                searchedValue = row[work_codice] 
                if flag_type == "list":
                    result_ = self.findCell_dataframe_RIV(df_rivisto, searchedValue, mapping_key, configurazione_rivisto[element])
                elif flag_type == "string":
                    result_ = self.findCell_dataframe_RIV_string(df_rivisto, searchedValue, mapping_key, configurazione_rivisto[element])
                
                if result_ == -2:
                    error_dict["error_ck_"+element].append(str(int(index)+2))
                    print("Coppia prestazione/agenda non trovata in rivisto: " +element)
                    out_message = "__> {}".format("Coppia prestazione/agenda non trovata in rivisto")
                    if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                        sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message 
                    else:
                        sheet[self.work_alert_column+str(int(index)+2)] = out_message
                elif isinstance(result_, str):
                    error_dict["error_ck_reverse_"+element].append(str(int(index)+2))
                    print("trovato errore su "+element+ " : " + result_)
                    out_message = "__> {}".format("Corrispondenza "+element+" non trovata in rivisto")
                    if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                        sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message 
                    else:
                        sheet[self.work_alert_column+str(int(index)+2)] = out_message

                    if sheet[self.work_map_value_column+str(int(index)+2)].value is not None:
                        sheet[self.work_map_value_column+str(int(index)+2)] = str(sheet[self.work_map_value_column+str(int(index)+2)].value) + "; \n" + result_ 
                    else:
                        sheet[self.work_map_value_column+str(int(index)+2)] = result_
                elif result_ == -1:
                    
                    print("trovata corrispondenza "+element)
                    '''out_message = "__> {}".format(element + " corrisponde in rivisto")
                    if sheet[self.work_alert_column+str(int(index)+2)].value is not None:
                        sheet[self.work_alert_column+str(int(index)+2)] = str(sheet[self.work_alert_column+str(int(index)+2)].value) + "; \n" + out_message #modificare colonna alert
                    else:
                        sheet[self.work_alert_column+str(int(index)+2)] = out_message'''
                
        xfile.save(self.file_data)     
        return error_dict

    