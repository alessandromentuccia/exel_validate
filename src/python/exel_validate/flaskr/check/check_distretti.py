import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import yaml
import logging
import re
import openpyxl
from flaskr.check_action import Check_action

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

class Check_distretti(Check_action):

    output_message = ""
    error_list = {}

    def __init__(self, file):
        #pass
        super().__init__(file)
        self.file = file
    
    def ck_distretti_inprestazione(self,error_dict):
        print("start checking if distretti are correct")
        error_dict.update({'error_distretti_inprestazione': []})
        distretto_dict_error = {}
        prestazioni_dict = {}

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet]#recupero sheet excel

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                Distretto_string = row[self.file.work_codice_distretto].split(self.file.work_delimiter)

                siss = ""
                siss_flag = False
                if Distretto_string is not None:
                    cod_pre_siss = str(row[self.file.work_codice_prestazione_siss]).strip()
                    for distretto in Distretto_string:
                        distretto = distretto.strip()
                        if distretto != "":
                            short_sheet = self.file.sheet_Distretti.loc[self.file.sheet_Distretti["Codice Distretto"] == distretto] #filtro catalogo sul codice distretto                     
                            
                            #print("disciplina " + str(disciplina["Cod Disciplina"]) + " " + str(disciplina["Codice Quesito"]))
                            print("prestazione del distretto in mapping 11:" + cod_pre_siss + " + " + distretto)
                            
                            if cod_pre_siss not in short_sheet["Codice SISS"].values:
                                siss_flag = True
                                prestazioni_dict[str(int(index)+2)] = cod_pre_siss
                                print("error distretto on index 1:" + str(int(index)+2))
                                distretto_dict_error = self.update_list_in_dict(distretto_dict_error, str(int(index)+2), distretto)
                            else:
                                if cod_pre_siss != siss and siss != "":
                                    print("error distretto on index 2:" + str(int(index)+2))
                                    siss_flag = True
                                    prestazioni_dict[str(int(index)+2)] = cod_pre_siss
                                    distretto_dict_error = self.update_list_in_dict(distretto_dict_error, str(int(index)+2), distretto)
                    siss = cod_pre_siss
                if siss_flag is True: #se durante il mapping con la sua prestazione, questa non viene rilevata, allora è errore
                    error_dict['error_distretti_inprestazione'].append(str(int(index)+2))
                

        out1 = ""
        out_message = ""
        for ind in error_dict['error_distretti_inprestazione']:
            out1 = out1 + "at index: " + ind + ", on distretti: " + ", ".join(distretto_dict_error[ind]) + ", \n"
            out_message = "__> Distretti: '{}' non previsti per la prestazione: '{}'".format(", ".join(distretto_dict_error[ind]), prestazioni_dict[ind])
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        self.output_message = self.output_message + "\nerror_distretti_inprestazione: \n" + out1
        
        xfile.save(self.file.file_data)
        return error_dict

    def ck_distretti_sintassi(self, error_dict):
        print("start checking if there is ',' separator between each distretti defined")
        error_dict.update({
            'error_distretti_caratteri_non_consentiti': [],
            'error_distretti_trovato_spazio': []
        })
        string_check = re.compile('1234567890,D;')

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.file.work_sheet) #recupero sheet excel

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                #print("Distretto: " + row["Codice Distretto"])
                if row[self.file.work_codice_distretto].strip() is not None:
                    r = row[self.file.work_codice_distretto].strip()
                    if(string_check.search(row[self.file.work_codice_distretto]) != None):
                        print("String contains other Characters.")
                        error_dict['error_distretti_caratteri_non_consentiti'].append(str(int(index)+2))
                    elif " " in r: 
                        print("string contain space")
                        error_dict['error_distretti_trovato_spazio'].append(str(int(index)+2))
    
        out1 = ", \n".join(error_dict['error_distretti_caratteri_non_consentiti'])
        self.output_message = self.output_message + "\nerror_distretti_caratteri_non_consentiti: \n" + "at index: \n" + out1
        out2 = ", \n".join(error_dict['error_distretti_trovato_spazio'])
        self.output_message = self.output_message + "\nerror_distretti_trovato_spazio: \n" + "at index: \n" + out2
        
        out_message = ""
        for ind in error_dict['error_distretti_caratteri_non_consentiti']:
            out_message = "__> Distretti presentano errori di sintassi: rilevati caratteri non consentiti"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        for ind in error_dict['error_distretti_trovato_spazio']:
            out_message = "__> Distretti presentano errori di sintassi: rilevati degli spazi non consentiti nella cella"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)

        return error_dict

    def ck_distretti_descrizione(self, error_dict):
        print("start checking if distretti have the correct description")
        error_dict.update({'error_distretti_descrizione': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.file.work_sheet) #recupero sheet excel

        distretti_dict_error = {}

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                distretto_string = row[self.file.work_codice_distretto].split(self.file.work_delimiter)
                description_list = row[self.file.work_descrizione_distretto].split(self.file.work_delimiter)
                flag_error = False
                '''if len(Distretto_string) != len(Description_list):
                    print("il numero di descrizioni è diverso dal numero di distretti all'indice " + str(index))
                    flag_error = True'''

                if distretto_string is not None:
                    for distretto in distretto_string:
                        distretto = distretto.strip()
                        if distretto != "":
                            distretto_catalogo = self.file.sheet_Distretti.loc[self.file.sheet_Distretti["Codice Distretto"] == distretto]                    
                            
                            try:
                                if distretto_catalogo["Distretti"].values[0] not in description_list:
                                    print("la descrizione distretto non è presente all'indice " + str(int(index)+2))
                                    flag_error = True
                                    distretti_dict_error = self.update_list_in_dict(distretti_dict_error, str(int(index)+2), distretto)
                            except:
                                if distretto_catalogo.size == 0:
                                    print("print distretto_catalogo2:" + distretto_catalogo)
                                    print("controllare manualmente qual'è il problema")
                                    logging.error("controllare manualmente il distretto: " + distretto + " all'indice: " + str(int(index)+2))
                                    flag_error = True
                                    distretti_dict_error = self.update_list_in_dict(distretti_dict_error, str(int(index)+2), distretto)
                if flag_error is True:
                    error_dict['error_distretti_descrizione'].append(str(int(index)+2))

        out1 = ""
        out_message = ""
        for ind in error_dict['error_distretti_descrizione']:
            out1 = out1 + "at index: " + ind + ", on distretto: " + ", ".join(distretti_dict_error[ind]) + ", \n"
            out_message = "__> Distretti: '{}' presentano errori nella descrizione".format(", ".join(distretti_dict_error[ind]))
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message #modificare colonna alert
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        self.output_message = self.output_message + "\nerror_distretti_descrizione: \n" + out1

        xfile.save(self.file.file_data)
        return error_dict

    def ck_distretti_operatori_logici(self, error_dict):
        print("start checking if there are the same logic op. for each prestazione")
        error_dict.update({
            'error_distretti_operatori_logici_non_necessari': [],
            'error_distretti_operatori_logici_mancante': [] })
        

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.file.work_sheet) #recupero sheet excel

        
        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                if row[self.file.work_operatore_logico_distretto] is not "" and row[self.file.work_codice_distretto].strip() is "":
                    error_dict['error_distretti_operatori_logici_non_necessari'].append(str(int(index)+2))

                elif row[self.file.work_operatore_logico_distretto] is "" and row[self.file.work_codice_distretto].strip() is not "":
                    error_dict['error_distretti_operatori_logici_mancante'].append(str(int(index)+2))
                    
        out1 = ""
        out_message = ""
        for ind in error_dict['error_distretti_operatori_logici_non_necessari']:
            out1 = out1 + "at index: " + ind
            out_message = "__> Operatore logico non necessario, non c'è codice distretto"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        self.output_message = self.output_message + "\nerror_distretti_operatori_logici: \n" + "at index: \n" + out1
        
        for ind in error_dict['error_distretti_operatori_logici_mancante']:
            out_message = "__> Distretti con Operatore logico assente"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message
            else:
                sheet[self.file.work_alert_column+ind] = out_message
        out2 = ", \n".join(error_dict['error_distretti_operatori_logici_mancante'])
        self.output_message = self.output_message + "\nerror_distretti_operatori_logici_mancante: \n" + "at index: \n" + out2
        
        xfile.save(self.file.file_data)
        return error_dict   