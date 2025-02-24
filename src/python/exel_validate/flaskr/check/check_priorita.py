import pandas as pd
import numpy as np
import requests
import xlrd
from xlsxwriter.utility import xl_rowcol_to_cell
import yaml
import logging
import re
import openpyxl
from flaskr.check_action import Check_action


logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

class Check_priorita(Check_action):

    output_message = ""
    error_list = {}

    def __init__(self, file):
        super().__init__(file)
        self.file = file
    
    def ck_prime_visite(self, error_dict):
        #Descrizione Prestazione SISS
        print("start checking if prestazione prime visite is correct")
        error_dict.update({'error_prime_visite': []})
        
        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile.get_sheet_by_name(self.file.work_sheet) #recupero sheet excel

        str_check = "PRIMA VISITA"
        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                if str_check in row[self.file.work_descrizione_prestazione_siss]:
                    logging.info("Prestazione PRIMA VISITA da controllare all'indice: " + str(int(index)+2))
                    if row[self.file.work_priorita_U] == "N" and row[self.file.work_priorita_primo_accesso_D] == "N" and row[self.file.work_priorita_primo_accesso_P] == "N" and row[self.file.work_priorita_primo_accesso_B] == "N": 
                        logging.error("trovato anomalia in check prime visite all'indice: " + str(int(index)+2))
                        error_dict["error_prime_visite"].append(str(int(index)+2))
        
        out1 = ", \n".join(error_dict['error_prime_visite'])
        self.output_message = self.output_message + "\nerror_prime_visite: \n" + "at index: \n" + out1
        
        out_message = ""
        for ind in error_dict['error_prime_visite']:
            out_message = "__> Rilevato errore di priorità per prestazione PRIMA VISITA"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message 
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)
        return error_dict

    '''Nel caso di prestazione visita di controllo, verificare se è presente il campo Accesso programmabile ZP'''
    def ck_controlli(self, error_dict):
        print("start checking if prestazione controlli is correct")
        error_dict.update({'error_controlli': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        str_check = "CONTROLLO"
        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                if str_check in row[self.file.work_descrizione_prestazione_siss]:
                    logging.info("Prestazione CONTROLLO da controllare all'indice: " + str(int(index)+2))
                    if row[self.file.work_accesso_programmabile_ZP] == "N": 
                        logging.error("trovato anomalia in check prestazione controllo all'indice: " + str(int(index)+2))
                        error_dict["error_controlli"].append(str(int(index)+2))
        
        out1 = ", \n".join(error_dict['error_controlli'])
        self.output_message = self.output_message + "\nerror_controlli: \n" + "at index: \n" + out1
        
        out_message = ""
        for ind in error_dict['error_controlli']:
            out_message = "__> Rilevato possibile errore di priorità per prestazione DI CONTROLLO"
            out_message = out_message + "\n _> controllare che l'accesso programmabile ZP non sia a N"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message 
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)
        return error_dict

    '''Nel caso di prestazioni per esami strumentale, controllare se le priorità sono definite'''
    def ck_esami_strumentali(self, error_dict):
        print("start checking if prestazione esami is correct")
        error_dict.update({'error_esami_strumentali': []})

        xfile = openpyxl.load_workbook(self.file.file_data) #recupero file excel da file system
        sheet = xfile[self.file.work_sheet] #recupero sheet excel

        str_check = "VISITA"

        for index, row in self.file.df_mapping.iterrows():
            if row[self.file.work_abilitazione_esposizione_siss] == "S":
                if str_check not in row[self.file.work_descrizione_prestazione_siss]:  
                    if row[self.file.work_priorita_U] == "N" and row[self.file.work_priorita_primo_accesso_D] == "N" and row[self.file.work_priorita_primo_accesso_P] == "N" and row[self.file.work_priorita_primo_accesso_B] == "N":
                        logging.info("Prestazione ESAME da controllare all'indice: " + str(int(index)+2))
                        if row[self.file.work_accesso_programmabile_ZP] == "N": 
                            logging.error("trovato anomalia in check esami strumentali all'indice: " + str(int(index)+2))
                            error_dict["error_esami_strumentali"].append(str(int(index)+2))
        
        out1 = ", \n".join(error_dict['error_esami_strumentali'])
        self.output_message = self.output_message + "\nerror_esami_strumentali: \n" + "at index: \n" + out1
        
        out_message = ""
        for ind in error_dict['error_esami_strumentali']:
            out_message = "__> Rilevato errore di priorità per prestazione di tipo esami strumentali"
            if sheet[self.file.work_alert_column+ind].value is not None:
                sheet[self.file.work_alert_column+ind] = str(sheet[self.file.work_alert_column+ind].value) + "; \n" + out_message 
            else:
                sheet[self.file.work_alert_column+ind] = out_message

        xfile.save(self.file.file_data)
        return error_dict