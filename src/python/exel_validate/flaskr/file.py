import os
import json
#from distutils.log import error
import logging
from collections import OrderedDict
#from functools import reduce
#from pathlib import Path
#from typing import Dict, List
import pandas as pd
import numpy as np
#import xlrd
import openpyxl
#from xlsxwriter.utility import xl_rowcol_to_cell
from dotenv import load_dotenv

#from flaskr.check.check_QD import Check_QD
#from flaskr.check.check_metodiche import Check_metodiche 
from flaskr.check.check_distretti import Check_distretti
from flaskr.check.check_priorita import Check_priorita
from flaskr.check.check_prestazione import Check_prestazione
from flaskr.check.check_canali import Check_canali
from flaskr.check.check_agende_interne import Check_agende_interne
from flaskr.check.check_inviante import Check_inviante
from flaskr.report_creation.rep_creation import Report_Creation #rep_creation.

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
f_handler = logging.FileHandler('generator.log', 'a+', 'utf-8')
c_handler = logging.StreamHandler()
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
f_handler.setFormatter(formatter)
c_handler.setFormatter(formatter)
logger.addHandler(f_handler)
logger.addHandler(c_handler)

RESULT_VALIDATION = "..\check_excel_result.txt"
load_dotenv()
CAT_NAME = os.getenv("CAT_SISS")
ROOT_DIR = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))



class File():

    controls_setted = {}
    
    validation_results = ""
    output_message = ""
    print("ROOT_DIR: " + ROOT_DIR+"\n")
    print("CAT_NAME: " + str(CAT_NAME)+"\n")
    catalogo_dir = os.path.join(ROOT_DIR, CAT_NAME) #ottengo directory catalogo attributi


    def __init__(self, data, excel_file, checked_dict):
        #self.output_message = ""
        #with open("./flaskr/config_validator.yml", "rt", encoding='utf8') as yamlfile:
        #    data = yaml.load(yamlfile, Loader=yaml.FullLoader)
        #logger.debug(data)
        
        
        #ottengo i singoli cataloghi QD, metodiche e distretti
        self.sheet_QD, self.sheet_Metodiche, self.sheet_Distretti = self.get_catalogo()

        self.work_sheet = data[0]["work_column"]["work_sheet"]
        self.work_N1 = data[0]["work_column"]["work_N1"]
        self.work_N2 = data[0]["work_column"]["work_N2"]
        self.work_descrizione_N1 = data[0]["work_column"]["work_descrizione_N1"]
        self.work_descrizione_N2 = data[0]["work_column"]["work_descrizione_N2"]
        self.work_codice_prestazione_siss = data[0]["work_column"]["work_codice_prestazione_siss"]
        self.work_descrizione_prestazione_siss = data[0]["work_column"]["work_descrizione_prestazione_siss"]
        self.work_codice_agenda_siss = data[0]["work_column"]["work_codice_agenda_siss"]
        #self.work_casi_1_n = data[0]["work_column"]["work_casi_1_n"]
        self.work_abilitazione_esposizione_siss = data[0]["work_column"]["work_abilitazione_esposizione_siss"]
        self.work_prenotabile_siss = data[0]["work_column"]["work_prenotabile_siss"]
        self.work_codici_disciplina_catalogo = data[0]["work_column"]["work_codici_disciplina_catalogo"]
        self.work_descrizione_disciplina_catalogo = data[0]["work_column"]["work_descrizione_disciplina_catalogo"]
        self.work_codice_QD = data[0]["work_column"]["work_codice_QD"]
        self.work_descrizione_QD = data[0]["work_column"]["work_descrizione_QD"]
        self.work_operatore_logico_QD = data[0]["work_column"]["work_operatore_logico_QD"]
        self.work_codice_metodica = data[0]["work_column"]["work_codice_metodica"]
        self.work_descrizione_metodica = data[0]["work_column"]["work_descrizione_metodica"]
        self.work_codice_distretto = data[0]["work_column"]["work_codice_distretto"]
        self.work_descrizione_distretto = data[0]["work_column"]["work_descrizione_distretto"]
        self.work_operatore_logico_distretto = data[0]["work_column"]["work_operatore_logico_distretto"]
        self.work_priorita_U = data[0]["work_column"]["work_priorita_U"]
        self.work_priorita_primo_accesso_D = data[0]["work_column"]["work_priorita_primo_accesso_D"]
        self.work_priorita_primo_accesso_P = data[0]["work_column"]["work_priorita_primo_accesso_P"]
        self.work_priorita_primo_accesso_B = data[0]["work_column"]["work_priorita_primo_accesso_B"]
        self.work_accesso_programmabile_ZP = data[0]["work_column"]["work_accesso_programmabile_ZP"]
        self.work_combinata = data[0]["work_column"]["work_combinata"]
        self.work_codice_agenda_interno = data[0]["work_column"]["work_codice_agenda_interno"]
        self.work_codice_prestazione_interno = data[0]["work_column"]["work_codice_prestazione_interno"]
        self.work_inviante = data[0]["work_column"]["work_inviante"]
        self.work_accesso_farmacia = data[0]["work_column"]["work_accesso_farmacia"]
        self.work_accesso_CCR = data[0]["work_column"]["work_accesso_CCR"]
        self.work_accesso_cittadino = data[0]["work_column"]["work_accesso_cittadino"]
        self.work_accesso_MMG = data[0]["work_column"]["work_accesso_MMG"]
        self.work_accesso_amministrativo = data[0]["work_column"]["work_accesso_amministrativo"]
        self.work_accesso_PAI = data[0]["work_column"]["work_accesso_PAI"]
        self.work_gg_preparazione = data[0]["work_column"]["work_gg_preparazione"]
        self.work_gg_refertazione = data[0]["work_column"]["work_gg_refertazione"]
        self.work_nota_operatore = data[0]["work_column"]["work_nota_operatore"]
        self.work_nota_agenda = data[0]["work_column"]["work_nota_agenda"]
        self.work_nota_revoca = data[0]["work_column"]["work_nota_revoca"]
        self.work_disciplina = data[0]["work_column"]["work_disciplina"]
        self.work_sesso = data[0]["work_column"]["work_sesso"]
        self.work_eta_min = data[0]["work_column"]["work_eta_min"]
        self.work_eta_max = data[0]["work_column"]["work_eta_max"]

        self.work_alert_column = data[1]["work_index"]["work_alert_column"]
        try:
            self.work_delimiter = data[2]["work_separator"]["work_delimiter"]
        except:
            self.work_delimiter = "," #valore di default
        self.file_data = excel_file

        self.df_mapping = pd.read_excel(self.file_data, sheet_name=self.work_sheet, converters={self.work_codici_disciplina_catalogo: str, self.work_codice_prestazione_siss: str, self.work_codice_metodica: str, self.work_codice_distretto: str}).replace(np.nan, '', regex=True)
        
        #print(checked_dict)
        self.controls_setted = checked_dict #checklist dei test da effettuare
        #self.analizer(df_mapping)


    def findCell(self, sh, searchedValue, start_col):
        result_coord = []
        result_value = []

        for row in range(sh.nrows):
            for col in range(start_col, start_col+1):
                myCell = sh.cell(row, col)
                myValue = sh.cell(row, self.work_index_codice_prestazione_SISS)
                abilita = sh.cell(row, self.work_index_abilitazione_esposizione_SISS)
                if myCell.value == searchedValue and abilita.value == "S":
                    result_coord.append(str(row) + "#" + str(col))
                    result_value.append(myValue.value)
                    #return row, col#xl_rowcol_to_cell(row, col)

        if result_coord == []:
            return -1
        return result_coord#, result_value

    '''def findCell_agenda(self, sh, searchedValue, start_col):
        result_coord = []
        result_value = []

        for row in range(sh.nrows):
            for col in range(start_col, start_col+1):
                myCell = sh.cell(row, col)
                myValue = sh.cell(row, self.work_index_codice_SISS_agenda) #Codice SISS agenda 15
                #abilita = sh.cell(row, self.work_index_abilitazione_esposizione_SISS-1) #abilitazione esposizione SISS 28
                if myCell.value == searchedValue: # and abilita.value == "S":
                    result_coord.append(str(row) + "|" + str(col))
                    result_value.append(myValue.value)
                    #return row, col#xl_rowcol_to_cell(row, col)

        if result_coord == []:
            return -1
        return result_coord#, result_value'''
                                    
    '''def findCell_agenda_II(self, df, searchedValue, name_column):#sheet_mapping, searchedAgenda, self.work_index_codice_SISS_agenda
        result_coord = []

        for rowIndex, row in df.iterrows(): #iterate over rows
            myCell = row[name_column]
            abilita = row[self.work_abilitazione_esposizione_siss]
            if myCell == searchedValue and abilita == "S":
                result_coord.append(str(rowIndex))

        if result_coord == []:
            return -1
        return result_coord'''

    '''Metodo che aggiunge elemento in una lista esistente o crea la lista nel caso
    non fosse presente'''
    def update_list_in_dict(self, dictio, index, element):
        if index in dictio.keys():
            dictio[index].append(element)
        else:
            dictio[index] = [element]
        return dictio

    '''Metodo per controllare il tipo di una colonna excel'''
    def column_validator(self):
        return ""

    def list_duplicates(self, seq):
        seen = set()
        seen_add = seen.add
        # adds all elements it doesn't know yet to seen and all other to seen_twice
        seen_twice = set( x for x in seq if x in seen or seen_add(x) )
        # turn the set into a list (as requested)
        return list( seen_twice )
    
    def get_catalogo(self):

        sheet_QD = pd.read_excel(self.catalogo_dir, sheet_name='QD', converters={"Cod Disciplina": str})
        print("sheet_QD caricato\n")
        #print(sheet_QD)
        sheet_Metodiche = pd.read_excel(self.catalogo_dir, sheet_name='METODICHE', converters={"Codice SISS": str, "Codice Metodica": str})
        print("sheet_Metodiche caricato\n")
        #print(sheet_Metodiche)
        sheet_Distretti = pd.read_excel(self.catalogo_dir, sheet_name='DISTRETTI', converters={"Codice SISS": str, "Codice Distretto": str})
        print("sheet_Distretti caricato\n")
        #print(sheet_Distretti)
        
        return sheet_QD, sheet_Metodiche, sheet_Distretti